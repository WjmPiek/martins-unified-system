from datetime import datetime, timedelta, date
from collections import defaultdict
from functools import wraps
import re
import secrets
import string
from difflib import SequenceMatcher
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file, current_app
from flask_login import login_required, current_user
from app.extensions import db
from sqlalchemy import text
from app.models import User, Role, Permission, AuditLog, Franchise, RoyaltyScale, MonthlyFigure, ImportJob, LiveEvent, LiveNotification, PerformancePageCache, ImportJobLog, WorkerHeartbeat, SystemEvent, EventSubscription, EventProcessingLog, RoyaltyGrowthProfile, RoyaltyAgreementProfile, RoyaltyCalculationSnapshot, RoyaltyOverride, FranchiseHealthSnapshot, BusinessInsight, InsightNarrative, WorkflowDefinition, WorkflowInstance, WorkflowStep, BusinessRule, EnterpriseTask, EnterpriseNotification, ScheduledJobDefinition, EnterpriseAuditTimeline, user_franchises
from app.franchise_context import set_selected_franchise
from app.permissions import MODULES, ACTIONS, ROLE_TEMPLATES, ROLE_DEFAULTS, permission_code
from app.audit import log_action
from app.performance.cache import cache_stats, invalidate_performance_cache
from app.performance.service import auto_hide_inactive_franchises, inactive_franchise_candidates, reactivate_franchise_performance, has_recent_performance_data, warm_performance_cache_for_period

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

PROTECTED_ADMIN_EMAIL = "wjm@martinsdirect.com"

ADMIN_SIDE_ROLE_NAMES = {"Admin", "Finance Manager", "Finance Assistant", "Regional Manager"}
FRANCHISE_SIDE_ROLE_NAMES = {"Franchise User", "Franchise Manager", "Franchise Employee", "Franchise Agent", "Read Only User"}
ADMIN_CREATABLE_ROLE_NAMES = ["Finance Manager", "Finance Assistant", "Regional Manager", "Franchise User"]
FRANCHISE_CREATABLE_ROLE_NAMES = ["Franchise Manager", "Franchise Employee", "Franchise Agent"]

ROLE_HELP_TEXT = {
    "Finance Manager": "Martins Funerals South Africa user. Sees the whole financial system and is not linked to one franchise.",
    "Finance Assistant": "Martins Funerals South Africa user. Finance support access and is not linked to one franchise.",
    "Regional Manager": "Martins Funerals South Africa user. Must be linked to the franchises/region they manage.",
    "Franchise User": "Franchise owner/user. Must be linked to the franchise(s) they own or operate.",
}
FINANCE_ADMIN_USERS = {
    "renette@martinsdirect.com": "Finance Manager",
    "lowhann@martinsdirect.com": "Finance Assistant",
    "lowhaan@martinsdirect.com": "Finance Assistant",
    "deon@martinsdirect.com": "Finance Assistant",
}


def ensure_user_hierarchy_roles():
    """Ensure the mother-company and franchise-level roles exist for the create-user screens."""
    descriptions = {
        "Finance Manager": "Martins Funerals South Africa finance manager",
        "Finance Assistant": "Martins Funerals South Africa finance assistant",
        "Regional Manager": "Martins regional manager linked to selected franchises",
        "Franchise User": "Franchise owner/user linked to selected franchise data",
        "Franchise Manager": "Manager created by a franchise user",
        "Franchise Employee": "Employee created by a franchise user",
        "Franchise Agent": "Agent created by a franchise user",
    }
    changed = False
    for role_name, description in descriptions.items():
        role = Role.query.filter_by(name=role_name).first()
        if not role:
            db.session.add(Role(name=role_name, description=description, is_system_role=True))
            changed = True
    if changed:
        db.session.flush()


def admin_creatable_roles():
    """Roles that Martins admin/finance users may create from the Admin Users page."""
    if is_current_user_admin():
        allowed = ADMIN_CREATABLE_ROLE_NAMES
    elif current_user.has_role("Finance Manager"):
        allowed = ["Finance Assistant", "Regional Manager", "Franchise User"]
    else:
        allowed = []
    return Role.query.filter(Role.name.in_(allowed)).order_by(Role.name).all()


def can_create_admin_user():
    # Admin must always be able to create Martins mother-company users.
    # Finance Manager may create allowed roles when Users Add permission is ticked.
    return is_current_user_admin() or (current_user.has_role("Finance Manager") and current_user.has_permission("users:add"))



def current_user_role_names():
    """Return role names robustly, including legacy/display role fields.

    Some older users show Admin in the UI but may not always have a populated
    role relationship in the current request. Treat the protected Martins admin
    account as Admin and read legacy single-role fields when present.
    """
    names = {role.name for role in getattr(current_user, "roles", []) or [] if getattr(role, "name", None)}

    for attr in ("role", "role_name", "user_role", "primary_role_name"):
        value = getattr(current_user, attr, None)
        if callable(value):
            try:
                value = value()
            except TypeError:
                value = None
        if value:
            names.add(str(value))

    email = (getattr(current_user, "email", "") or "").lower()
    full_name = (getattr(current_user, "full_name", "") or "").lower()
    if email == PROTECTED_ADMIN_EMAIL or full_name == "wjm piek":
        names.add("Admin")

    return names


def is_current_user_admin():
    return bool(current_user_role_names() & {"Admin", "Super Admin"})

def is_current_user_finance_import_user():
    return bool(current_user_role_names() & {"Finance Manager", "Finance Assistant"})


def can_view_imports_data():
    # Imports & Data is visible only to Admin and finance import users.
    # Admin sees all import tools. Finance Manager/Assistant see only Monthly Figures PDF.
    return is_current_user_admin() or is_current_user_finance_import_user()


def user_role_names(user):
    return {role.name for role in user.roles}


def is_admin_side_user(user):
    names = user_role_names(user)
    email = (user.email or "").lower()
    return email in FINANCE_ADMIN_USERS or bool(names & ADMIN_SIDE_ROLE_NAMES)


def is_franchise_side_user(user):
    names = user_role_names(user)
    return bool(names & FRANCHISE_SIDE_ROLE_NAMES) and not is_admin_side_user(user)


def is_role_admin_side(role_name):
    return role_name in ADMIN_SIDE_ROLE_NAMES


def role_requires_franchise_scope(role_name):
    return role_name in {"Regional Manager", "Franchise User"}


def normalise_user_scope_for_role(user, role_name, franchise_ids=None):
    """Keep admin-side users out of franchise hierarchy and scope franchise roles correctly."""
    user.parent_franchise_user_id = None

    if role_requires_franchise_scope(role_name):
        selected_franchises = Franchise.query.filter(
            Franchise.id.in_(franchise_ids or []),
            Franchise.is_performance_active == True,
        ).order_by(Franchise.business_name).all()
        if not selected_franchises:
            return False, "Please link at least one active franchise for Regional Manager or Franchise User accounts."
        user.assigned_franchises = selected_franchises
        return True, ""

    # Finance/Admin-side users are Martins users. Finance Manager is linked to all active franchises.
    if is_role_admin_side(role_name):
        if role_name == "Finance Manager":
            user.assigned_franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
        else:
            user.assigned_franchises = []
        return True, ""

    # Unknown roles are not allowed from the Admin create form.
    return False, "Please select a valid Admin-created role."


def tidy_finance_admin_users():
    # These named people are Martins Funerals South Africa/admin-side users, not franchise users.
    # Remove franchise links and accidental Franchise User/Manager roles, then keep their admin-side finance role.
    changed = 0
    for email, role_name in FINANCE_ADMIN_USERS.items():
        user = User.query.filter(db.func.lower(User.email) == email).first()
        if not user:
            continue
        wanted_role = Role.query.filter_by(name=role_name).first()
        cleaned_roles = [role for role in user.roles if role.name not in FRANCHISE_SIDE_ROLE_NAMES]
        if wanted_role and wanted_role not in cleaned_roles:
            cleaned_roles.append(wanted_role)
        if set(cleaned_roles) != set(user.roles):
            user.roles = cleaned_roles
            changed += 1
        if role_name == "Finance Manager":
            active_franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
            if set(user.assigned_franchises) != set(active_franchises):
                user.assigned_franchises = active_franchises
                changed += 1
        elif user.assigned_franchises:
            user.assigned_franchises = []
            changed += 1
    return changed


def can_create_regional_manager():
    # Controlled by User Roles: grant Franchise Management Manage plus Users Add/Edit.
    return current_user.has_permission("franchise_management:manage") and (
        current_user.has_permission("users:add") or current_user.has_permission("users:edit")
    )


def can_assign_franchise_links():
    # Admin must always be able to link Regional Manager and Franchise User accounts.
    if is_current_user_admin() or current_user.has_role("Super Admin"):
        return True
    return current_user.has_permission("franchise_management:manage") and current_user.has_permission("users:edit")


def can_bulk_import_users():
    return current_user.has_permission("users:add") or current_user.has_permission("users:import")


def can_manage_old_franchises():
    names = current_user_role_names()
    return bool(names & {"Admin", "Super Admin", "Finance Manager"}) or current_user.has_permission("performance:manage_inactive")




def ordered_franchises_for_user(user):
    linked = list(getattr(user, "assigned_franchises", []) or [])
    if not linked:
        return []
    primary_id = db.session.execute(
        db.select(user_franchises.c.franchise_id)
        .where(user_franchises.c.user_id == user.id)
        .where(user_franchises.c.is_primary == True)
    ).scalar()
    linked_sorted = sorted(linked, key=lambda item: item.business_name or "")
    if primary_id:
        primary = [item for item in linked_sorted if item.id == primary_id]
        rest = [item for item in linked_sorted if item.id != primary_id]
        return primary + rest
    return linked_sorted

def active_linked_franchises_for_user(user):
    return [franchise for franchise in ordered_franchises_for_user(user) if getattr(franchise, "is_performance_active", True)]


def old_linked_franchises_for_user(user):
    return [franchise for franchise in ordered_franchises_for_user(user) if not getattr(franchise, "is_performance_active", True)]


def franchise_user_has_active_data(user):
    return bool(active_linked_franchises_for_user(user))


def franchise_user_has_recent_kpi_data(user, month=None, year=None):
    """A franchise user is listed only when at least one linked franchise has KPI data in the last 3 months."""
    now = datetime.utcnow()
    month = month or now.month
    year = year or now.year
    for franchise in active_linked_franchises_for_user(user):
        if has_recent_performance_data(franchise.id, month, year, 3):
            return True
    return False


def all_franchise_owner_users():
    query_users = User.query.order_by(User.name, User.surname).all()
    return [
        user for user in query_users
        if user.has_role("Franchise User") and not getattr(user, "parent_franchise_user_id", None)
    ]


def active_recent_franchise_owner_users(month=None, year=None):
    return [
        user for user in all_franchise_owner_users()
        if franchise_user_has_recent_kpi_data(user, month, year)
    ]


def repair_existing_user_visibility():
    """Repair legacy imported records so the admin pages do not hide them.

    Some users were imported before the Admin/Franchise/Employee split existed.
    They can have an email address in users, but no correct role/scope, which
    causes duplicate-email errors while the person is invisible on the pages.
    This keeps the account and makes it visible in the correct tab.
    """
    ensure_user_hierarchy_roles()
    changed = 0

    finance_role = Role.query.filter_by(name="Finance Assistant").first()
    for email in ("lowhann@martinsdirect.com", "lowhaan@martinsdirect.com"):
        user = User.query.filter(db.func.lower(User.email) == email).first()
        if user and finance_role and finance_role not in user.roles:
            user.roles = [role for role in user.roles if role.name not in FRANCHISE_SIDE_ROLE_NAMES]
            user.roles.append(finance_role)
            user.parent_franchise_user_id = None
            changed += 1

    franchise_role = Role.query.filter_by(name="Franchise User").first()
    david = User.query.filter(db.func.lower(User.email) == "david@martinsfunerals.co.za").first()
    if david and franchise_role and franchise_role not in david.roles and not any(role.name in ADMIN_SIDE_ROLE_NAMES for role in david.roles):
        david.roles = [role for role in david.roles if role.name not in {"Franchise Manager", "Franchise Employee", "Franchise Agent"}]
        david.roles.append(franchise_role)
        david.parent_franchise_user_id = None
        changed += 1

    employee_role_names = {"Franchise Manager", "Franchise Employee", "Franchise Agent"}
    owner_ids = {owner.id for owner in all_franchise_owner_users()}
    for user in User.query.order_by(User.id).all():
        if user.id in owner_ids or any(role.name in ADMIN_SIDE_ROLE_NAMES for role in user.roles):
            continue
        if any(role.name in employee_role_names for role in user.roles):
            if not user.parent_franchise_user_id and user.created_by_user_id in owner_ids:
                user.parent_franchise_user_id = user.created_by_user_id
                changed += 1
            continue
        if user.parent_franchise_user_id or (user.created_by_user_id in owner_ids):
            role = Role.query.filter_by(name="Franchise Employee").first()
            if role and role not in user.roles:
                user.roles.append(role)
                changed += 1
            if not user.parent_franchise_user_id and user.created_by_user_id in owner_ids:
                user.parent_franchise_user_id = user.created_by_user_id
                changed += 1

    if changed:
        db.session.flush()
    return changed


def is_protected_admin_user(user):
    return bool(user and (user.email or "").lower() == PROTECTED_ADMIN_EMAIL)



def can_change_user_roles():
    # Admin must always be able to correct user roles. Finance Assistant still
    # requires the Users Edit permission.
    names = user_role_names(current_user)
    if "Admin" in names or "Super Admin" in names:
        return True
    return "Finance Assistant" in names and current_user.has_permission("users:edit")


def clean_franchise_name(value):
    name = str(value or "").strip()
    if not name or name.upper() == "TOTAL":
        return ""
    return " ".join(name.split())


def slugify_email_part(value):
    value = value.lower().replace("&", "and")
    value = re.sub(r"\(f\)", "", value)
    value = re.sub(r"[^a-z0-9]+", ".", value)
    value = re.sub(r"\.+", ".", value).strip(".")
    return value or "franchise"


def temporary_password(length=14):
    alphabet = string.ascii_letters + string.digits + "!@#"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def get_or_create_role(role_name):
    role = Role.query.filter_by(name=role_name).first()
    if role:
        return role
    role = Role(name=role_name, description=f"Imported {role_name} role", is_system_role=True)
    db.session.add(role)
    db.session.flush()
    return role


def get_or_create_user(name, surname, email, role_name, franchises=None, password=None):
    email = email.strip().lower()
    user = User.query.filter_by(email=email).first()
    created = False
    if not user:
        user = User(name=name, surname=surname, email=email, is_active=True, is_active_account=True)
        user.set_password(password or temporary_password())
        db.session.add(user)
        db.session.flush()
        created = True
    else:
        user.name = user.name or name
        user.surname = user.surname or surname
        user.is_active = True
        user.is_active_account = True

    role = get_or_create_role(role_name)
    if role not in user.roles:
        user.roles.append(role)

    if franchises is not None:
        for franchise in franchises:
            if franchise not in user.assigned_franchises:
                user.assigned_franchises.append(franchise)
    return user, created


def permission_required(code):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Admin/Super Admin must never be blocked from Admin screens by
            # missing seeded permission rows or legacy user-role mismatches.
            if is_current_user_admin():
                return func(*args, **kwargs)
            if not current_user.has_permission(code):
                abort(403)
            return func(*args, **kwargs)
        return wrapper
    return decorator



def seed_permissions_and_roles():
    for module_index, module in enumerate(MODULES):
        for action_index, action in enumerate(ACTIONS):
            code = permission_code(module, action)
            permission = Permission.query.filter_by(code=code).first()
            if not permission:
                permission = Permission(
                    module=module,
                    action=action,
                    code=code,
                    label=f"{action.title()} {module}",
                    sort_order=(module_index * 100) + action_index,
                )
                db.session.add(permission)
    db.session.flush()

    all_permissions = Permission.query.all()
    permissions_by_code = {permission.code: permission for permission in all_permissions}

    for role_name, description in ROLE_TEMPLATES.items():
        role = Role.query.filter_by(name=role_name).first()
        if not role:
            role = Role(name=role_name, description=description, is_system_role=True)
            db.session.add(role)
            db.session.flush()
        defaults = ROLE_DEFAULTS.get(role_name, {})
        if defaults == "ALL":
            role.permissions = list(all_permissions)
        else:
            role.permissions = []
            for module, actions in defaults.items():
                for action in actions:
                    permission = permissions_by_code.get(permission_code(module, action))
                    if permission:
                        role.permissions.append(permission)
    db.session.commit()


@admin_bp.route("/seed")
@login_required
@permission_required("user_roles:manage")
def seed():
    seed_permissions_and_roles()
    flash("Default roles and permissions have been created/updated.", "success")
    return redirect(url_for("admin.roles"))


@admin_bp.route("/users")
@login_required
@permission_required("users:view")
def users():
    repair_existing_user_visibility()
    db.session.commit()
    # Keep the franchise selector clean: branches with no KPI data in the last 3 months
    # are hidden automatically and shown in the Old Franchises tab until reactivated.
    now = datetime.utcnow()
    auto_hide_inactive_franchises(now.month, now.year, [franchise.id for franchise in Franchise.query.all()], current_user.id)
    franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
    old_franchise_rows = inactive_franchise_candidates(now.month, now.year, [franchise.id for franchise in Franchise.query.order_by(Franchise.business_name).all()])
    old_franchises = [row for row in old_franchise_rows if not row["is_performance_active"]]
    selected_franchise_id = request.args.get("franchise_id", type=int)

    if selected_franchise_id:
        selected_franchise = Franchise.query.get_or_404(selected_franchise_id)
        if set_selected_franchise(selected_franchise.id, franchise_view_mode=True):
            flash(f"Opened {selected_franchise.business_name}. You can edit its details and royalty scale below.", "success")
            return redirect(url_for("franchise.details"))
        flash("You do not have access to that franchise.", "danger")
        return redirect(url_for("admin.users"))

    tidy_finance_admin_users()
    db.session.commit()

    all_users = User.query.order_by(User.name, User.surname).all()
    mother_company_users = [
        user for user in all_users
        if user.has_role("Admin")
        or user.has_role("Finance Manager")
        or user.has_role("Finance Assistant")
        or user.has_role("Regional Manager")
    ]
    franchise_owner_users = active_recent_franchise_owner_users(now.month, now.year)
    all_franchise_owner_user_rows = all_franchise_owner_users()
    franchise_employee_users = [
        user for user in all_users
        if getattr(user, "parent_franchise_user_id", None)
        or user.has_role("Franchise Manager")
        or user.has_role("Franchise Employee")
        or user.has_role("Franchise Agent")
    ]
    admin_side_users = mother_company_users
    franchise_side_users = franchise_owner_users
    all_franchise_side_users = franchise_owner_users + franchise_employee_users
    old_franchise_users = [
        user for user in all_franchise_owner_user_rows
        if ordered_franchises_for_user(user) and not franchise_user_has_recent_kpi_data(user, now.month, now.year)
    ]
    other_users = [user for user in all_users if user not in mother_company_users and user not in franchise_owner_users and user not in franchise_employee_users]
    linked_franchise_groups = []
    for user in franchise_owner_users:
        linked = ordered_franchises_for_user(user)
        if len(linked) > 1:
            linked_franchise_groups.append({"user": user, "main": linked[0], "franchises": linked})

    return render_template(
        "admin/users.html",
        users=all_users,
        admin_side_users=admin_side_users,
        franchise_side_users=franchise_side_users,
        mother_company_users=mother_company_users,
        franchise_owner_users=franchise_owner_users,
        franchise_employee_users=franchise_employee_users,
        other_users=other_users,
        old_franchise_users=old_franchise_users,
        linked_franchise_groups=linked_franchise_groups,
        roles=Role.query.order_by(Role.name).all(),
        franchises=franchises,
        old_franchises=old_franchises,
        selected_franchise=None,
        can_assign_franchise_links=can_assign_franchise_links(),
        can_change_user_roles=can_change_user_roles(),
        can_manage_old_franchises=can_manage_old_franchises(),
        admin_side_role_names=ADMIN_SIDE_ROLE_NAMES,
        franchise_side_role_names=FRANCHISE_SIDE_ROLE_NAMES,
        admin_creatable_roles=admin_creatable_roles(),
        can_create_admin_user=can_create_admin_user(),
        role_help_text=ROLE_HELP_TEXT,
    )


@admin_bp.route("/franchise-users")
@login_required
@permission_required("users:view")
def franchise_users():
    repair_existing_user_visibility()
    now = datetime.utcnow()
    auto_hide_inactive_franchises(now.month, now.year, [franchise.id for franchise in Franchise.query.all()], current_user.id)
    db.session.commit()
    franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
    franchise_users = all_franchise_owner_users()
    return render_template(
        "admin/franchise_users.html",
        franchise_users=franchise_users,
        roles=Role.query.filter(Role.name.in_(["Franchise User"])).order_by(Role.name).all(),
        franchises=franchises,
        can_assign_franchise_links=can_assign_franchise_links(),
    )


@admin_bp.route("/users/create", methods=["GET", "POST"])
@login_required
def create_admin_user():
    if not can_create_admin_user():
        abort(403)

    if request.method == "GET":
        franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
        return render_template(
            "admin/create_martins_user.html",
            admin_creatable_roles=admin_creatable_roles(),
            franchises=franchises,
        )

    name = request.form.get("name", "").strip()
    surname = request.form.get("surname", "").strip()
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "").strip()
    role_id = request.form.get("role_id", type=int)
    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids")]

    if not name or not surname or not email or not password or not role_id:
        flash("Name, surname, email, password and role are required.", "danger")
        return redirect(url_for("admin.users"))

    existing = User.query.filter(db.func.lower(User.email) == email).first()
    if existing:
        existing_roles = ", ".join(role.name for role in existing.roles) or "No role"
        flash(f"Email {email} already exists as {existing.full_name} ({existing_roles}). Use Edit on the correct page to set a password or activate the user.", "danger")
        return redirect(url_for("admin.users"))

    role = Role.query.get_or_404(role_id)
    allowed_roles = {item.name for item in admin_creatable_roles()}
    if role.name not in allowed_roles:
        flash("You are not allowed to create that user role.", "danger")
        return redirect(url_for("admin.users"))

    user = User(
        name=name,
        surname=surname,
        email=email,
        is_active=True,
        is_active_account=True,
        parent_franchise_user_id=None,
        created_by_user_id=current_user.id,
    )
    user.set_password(password)
    user.roles.append(role)

    ok, message = normalise_user_scope_for_role(user, role.name, franchise_ids)
    if not ok:
        flash(message, "danger")
        return redirect(url_for("admin.users"))

    db.session.add(user)
    db.session.commit()
    log_action("Users", "Created admin-managed user", f"User: {email}; Role: {role.name}")
    flash(f"User {user.full_name} was created as {role.name}.", "success")
    if role.name == "Franchise User":
        return redirect(url_for("admin.franchise_users"))
    return redirect(url_for("admin.users"))


@admin_bp.route("/franchises/<int:franchise_id>/reactivate-performance", methods=["POST"])
@login_required
@permission_required("users:view")
def reactivate_old_franchise(franchise_id):
    if not can_manage_old_franchises():
        abort(403)
    franchise = reactivate_franchise_performance(franchise_id, current_user.id)
    if franchise:
        log_action("Franchise", "Reactivated old franchise", f"Franchise: {franchise.business_name}; ID: {franchise.id}")
        flash(f"{franchise.business_name} has been activated again and will be included in details, targets, graphs and calculations.", "success")
        return redirect(url_for("admin.users", franchise_id=franchise.id))
    flash("Franchise could not be found.", "danger")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:user_id>/roles", methods=["POST"])
@login_required
def update_user_roles(user_id):
    if not can_change_user_roles():
        flash("Only Admin and Finance Assistant users with Users Edit permission may change user roles.", "danger")
        return redirect(url_for("admin.users"))

    user = User.query.get_or_404(user_id)

    if user.email and user.email.lower() == PROTECTED_ADMIN_EMAIL:
        admin_role = Role.query.filter_by(name="Admin").first()
        if admin_role and admin_role not in user.roles:
            user.roles.append(admin_role)
            db.session.commit()
        flash("Primary system administrator is protected. Roles cannot be changed.", "warning")
        return redirect(url_for("admin.users"))

    role_ids = [int(role_id) for role_id in request.form.getlist("role_ids")]
    selected_roles = Role.query.filter(Role.id.in_(role_ids)).all() if role_ids else []
    if any(role.name == "Regional Manager" for role in selected_roles) and not can_create_regional_manager():
        flash("Your role does not have permission to create or assign Regional Manager users.", "danger")
        return redirect(url_for("admin.users"))
    selected_role_names = {role.name for role in selected_roles}
    if "Admin" in selected_role_names and (user.email or "").lower() != PROTECTED_ADMIN_EMAIL:
        flash("The Admin role is locked to wjm@martinsdirect.com only.", "danger")
        return redirect(url_for("admin.users"))
    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids")]

    # Mother-company finance/admin users must never sit under a franchise. Finance Manager is linked to all active franchise users/franchises.
    if selected_role_names & {"Admin", "Finance Manager", "Finance Assistant"}:
        user.parent_franchise_user_id = None
        if "Finance Manager" in selected_role_names:
            user.assigned_franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
        else:
            user.assigned_franchises = []
    elif selected_role_names & {"Regional Manager", "Franchise User"}:
        user.parent_franchise_user_id = None
        selected_franchises = Franchise.query.filter(
            Franchise.id.in_(franchise_ids or []),
            Franchise.is_performance_active == True,
        ).order_by(Franchise.business_name).all()
        if not selected_franchises:
            flash("Regional Manager and Franchise User accounts must be linked to at least one active franchise.", "danger")
            return redirect(url_for("admin.users"))
        user.assigned_franchises = selected_franchises
    else:
        # Admin > Users is only for Martins users and registered franchise owner/user accounts.
        # Franchise employees are managed separately under Admin > Employees and created by franchise owners.
        user.parent_franchise_user_id = None

    user.roles = selected_roles
    log_action("Users", "Updated user roles and scope", f"User: {user.full_name}")
    db.session.commit()
    flash(f"User and scope updated for {user.full_name}.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:user_id>/update", methods=["POST"])
@login_required
@permission_required("users:edit")
def update_user(user_id):
    user = User.query.get_or_404(user_id)

    if is_protected_admin_user(user):
        flash("The primary Admin account is locked and cannot be edited.", "danger")
        return redirect(url_for("admin.users"))

    name = request.form.get("name", "").strip()
    surname = request.form.get("surname", "").strip()
    email = request.form.get("email", "").strip().lower()
    is_active = request.form.get("is_active") == "1"
    role_ids = [int(role_id) for role_id in request.form.getlist("role_ids") if str(role_id).isdigit()]
    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids") if str(item).isdigit()]

    if not name or not surname or not email:
        flash("Name, surname and email are required.", "danger")
        return redirect(url_for("admin.users"))

    duplicate = User.query.filter(db.func.lower(User.email) == email, User.id != user.id).first()
    if duplicate:
        flash("Another user already uses that email address.", "danger")
        return redirect(url_for("admin.users"))

    selected_roles = Role.query.filter(Role.id.in_(role_ids)).all() if role_ids else []
    if not selected_roles:
        flash("Please select at least one role.", "danger")
        return redirect(url_for("admin.users"))

    selected_role_names = {role.name for role in selected_roles}
    if "Admin" in selected_role_names and email != PROTECTED_ADMIN_EMAIL:
        flash("The Admin role is locked to wjm@martinsdirect.com only.", "danger")
        return redirect(url_for("admin.users"))
    if "Regional Manager" in selected_role_names and not can_create_regional_manager():
        flash("Your role does not have permission to assign Regional Manager users.", "danger")
        return redirect(url_for("admin.users"))

    user.name = name
    user.surname = surname
    user.email = email
    user.is_active = is_active
    user.is_active_account = is_active
    password = request.form.get("password", "").strip()
    if password:
        user.set_password(password)

    if selected_role_names & {"Admin", "Finance Manager", "Finance Assistant"}:
        user.parent_franchise_user_id = None
        if "Finance Manager" in selected_role_names:
            user.assigned_franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
        else:
            user.assigned_franchises = []
    elif selected_role_names & {"Regional Manager", "Franchise User"}:
        user.parent_franchise_user_id = None
        selected_franchises = Franchise.query.filter(
            Franchise.id.in_(franchise_ids or []),
            Franchise.is_performance_active == True,
        ).order_by(Franchise.business_name).all()
        if not selected_franchises:
            flash("Regional Manager and Franchise User accounts must be linked to at least one active franchise.", "danger")
            return redirect(url_for("admin.users"))
        user.assigned_franchises = selected_franchises
    else:
        user.parent_franchise_user_id = None

    user.roles = selected_roles
    log_action("Users", "Updated user details", f"User: {user.full_name}; Email: {user.email}")
    db.session.commit()
    flash(f"User updated for {user.full_name}.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@permission_required("users:delete")
def delete_user(user_id):
    user = User.query.get_or_404(user_id)

    if is_protected_admin_user(user):
        flash("The primary Admin account is locked and cannot be deleted or deactivated.", "danger")
        return redirect(url_for("admin.users"))

    user.is_active = False
    log_action("Users", "Deactivated user", f"User: {user.full_name}")
    db.session.commit()
    flash(f"{user.full_name} has been deactivated. User data has been kept.", "success")
    return redirect(url_for("admin.users"))



@admin_bp.route("/users/cleanup-inactive", methods=["POST"])
@login_required
@permission_required("users:delete")
def cleanup_inactive_users():
    cutoff = datetime.utcnow() - timedelta(days=60)
    users = User.query.filter(User.email != PROTECTED_ADMIN_EMAIL).all()
    count = 0

    for user in users:
        last_seen = getattr(user, "last_login_at", None) or getattr(user, "created_at", None)
        if user.is_active and last_seen and last_seen < cutoff:
            user.is_active = False
            count += 1

    db.session.commit()
    log_action("Users", "Deactivated inactive users", f"Count: {count}")
    flash(f"{count} inactive user(s) deactivated. Their data was kept.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/import-franchise-users", methods=["GET", "POST"])
@login_required
def import_franchise_users():
    if not is_current_user_admin():
        abort(403)

    allowed_roles = ["Franchise User", "Franchise Manager"]
    if can_create_regional_manager():
        allowed_roles.append("Regional Manager")

    if request.method == "POST":
        target_role = request.form.get("target_role", "Franchise User").strip()
        if target_role not in allowed_roles:
            flash("You are not allowed to create that user role.", "danger")
            return redirect(url_for("admin.import_franchise_users"))

        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Please upload the Excel file with franchise names in row 1.", "danger")
            return redirect(url_for("admin.import_franchise_users"))

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
            raw_names = [cell.value for cell in worksheet[1]]
        except Exception as exc:
            flash(f"Could not read Excel file: {exc}", "danger")
            return redirect(url_for("admin.import_franchise_users"))

        franchise_names = []
        seen = set()
        removed_totals = 0
        for value in raw_names:
            cleaned = clean_franchise_name(value)
            if not cleaned:
                if str(value or "").strip().upper() == "TOTAL":
                    removed_totals += 1
                continue
            key = cleaned.lower()
            if key not in seen:
                franchise_names.append(cleaned)
                seen.add(key)

        if not franchise_names:
            flash("No franchise names were found in row 1 after removing TOTAL columns.", "warning")
            return redirect(url_for("admin.import_franchise_users"))

        franchises = []
        franchises_created = 0
        users_created = 0
        users_updated = 0
        generated = []

        for franchise_name in franchise_names:
            franchise = Franchise.query.filter(db.func.lower(Franchise.business_name) == franchise_name.lower()).first()
            if not franchise:
                franchise = Franchise(business_name=franchise_name, franchise_code=slugify_email_part(franchise_name).upper()[:20])
                db.session.add(franchise)
                db.session.flush()
                franchises_created += 1
            franchises.append(franchise)

            email = f"{slugify_email_part(franchise_name)}@martinsdirect.com"
            password = temporary_password()
            user, created = get_or_create_user(franchise_name, "User", email, target_role, [], password)
            if created:
                users_created += 1
                generated.append((user.full_name, user.email, password, "Not linked - assign manually", target_role))
            else:
                users_updated += 1

        # Legacy behaviour removed: the import must not create Renette, Lowhaan or Deon as separate
        # Finance Manager / Finance Assistant users. Finance permissions now belong to the Admin role
        # or to roles explicitly assigned in User Roles.
        finance_created = 0
        finance_updated = 0
        db.session.commit()
        log_action(
            "Users",
            "Imported franchise users from Excel",
            f"Franchises: {len(franchises)}, new franchises: {franchises_created}, new users: {users_created}, updated users: {users_updated}, totals removed: {removed_totals}",
        )

        flash(
            f"Import complete. {len(franchises)} franchises processed, {franchises_created} new franchises created, "
            f"{users_created} franchise users created, {users_updated} users updated, {removed_totals} TOTAL columns removed. "
            "No separate finance users were created.",
            "success",
        )

        return render_template(
            "admin/import_franchise_users.html",
            allowed_roles=allowed_roles,
            selected_role=target_role,
            generated=generated,
            import_complete=True,
        )

    return render_template(
        "admin/import_franchise_users.html",
        allowed_roles=allowed_roles,
        selected_role="Franchise User",
        generated=[],
        import_complete=False,
    )



def normalize_franchise_key(value):
    """Create a forgiving key for matching spreadsheet branch names to Franchise records."""
    text = str(value or "").strip().lower()
    text = re.sub(r"\(f\)", "", text)
    text = text.replace("martin's", "martins")
    text = text.replace("martins funerals", "")
    text = text.replace("martins funeral", "")
    text = text.replace("martins begrafnisdienste", "")
    text = text.replace("begrafnisdienste", "")
    text = text.replace("funerals", "")
    text = re.sub(r"\bpty\b|\bltd\b|\blimited\b|\btas\b|\bt/a\b|\bck\b", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


AFRIKAANS_MONTHS = {
    "januarie": 1, "jan": 1,
    "februarie": 2, "feb": 2,
    "maart": 3, "mrt": 3,
    "april": 4, "apr": 4,
    "mei": 5,
    "junie": 6, "jun": 6,
    "julie": 7, "jul": 7,
    "augustus": 8, "aug": 8,
    "september": 9, "sep": 9,
    "oktober": 10, "okt": 10,
    "november": 11, "nov": 11,
    "desember": 12, "des": 12,
}


def parse_contract_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace("  ", " ")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = re.match(r"^(\d{1,2})\s+([A-Za-zÀ-ÿ]+)\s+(\d{4})$", text)
    if match:
        day = int(match.group(1))
        month = AFRIKAANS_MONTHS.get(match.group(2).lower())
        year = int(match.group(3))
        if month:
            return date(year, month, day)
    return None


def clean_excel_text(value):
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def unique_join(values):
    seen = set()
    result = []
    for value in values:
        text = clean_excel_text(value)
        if not text:
            continue
        key = text.lower()
        if key not in seen:
            result.append(text)
            seen.add(key)
    return "; ".join(result)


def last_non_empty(values):
    chosen = ""
    for value in values:
        text = clean_excel_text(value)
        if text:
            chosen = text
    return chosen


def last_contract_date(values):
    chosen = None
    for value in values:
        parsed = parse_contract_date(value)
        if parsed:
            chosen = parsed
    return chosen


def newest_contract_date(values):
    """Return the newest valid date from a set of spreadsheet cells.

    Contract summary imports are a full refresh for matched franchises.
    When the Excel sheet is updated and uploaded again, the database must be
    overwritten from the new file instead of keeping older dates.
    """
    parsed_dates = []
    for value in values:
        parsed = parse_contract_date(value)
        if parsed:
            parsed_dates.append(parsed)
    return max(parsed_dates) if parsed_dates else None


def set_auto_gross_method_from_agreement(franchise):
    franchise.royalty_gross_method = "new" if (franchise.agreement_start_date and franchise.agreement_start_date.year >= 2018) else "old"


def sync_royalty_scales_from_contract_file(franchise, parsed_rows, raw_scale_lines, minimum):
    """Fully sync royalty-scale fields from the latest uploaded Contract Summary file.

    This intentionally replaces the database values every time the Excel file is
    uploaded. It avoids the old behaviour where blank/changed cells left old
    values behind and made the Franchise Details page look unchanged.
    """
    franchise.imported_royalty_scale_text = "\n".join(raw_scale_lines or [])
    franchise.imported_royalty_percentage = (parsed_rows[0].get("percentage") if parsed_rows else 0) or 0
    franchise.minimum_royalty_amount = minimum if minimum is not None else 0

    RoyaltyScale.query.filter_by(franchise_id=franchise.id).delete()
    db.session.flush()
    for index, parsed in enumerate(parsed_rows or [], start=1):
        db.session.add(RoyaltyScale(
            franchise_id=franchise.id,
            row_number=index,
            amount_from=parsed.get("amount_from") or 0,
            amount_to=parsed.get("amount_to") or 999999999,
            percentage=parsed.get("percentage") or 0,
        ))
    return len(parsed_rows or [])


def parse_money_token(value):
    if value is None:
        return None
    text = str(value)
    text = re.sub(r"[^0-9.,-]", "", text).replace(" ", "")
    if not text:
        return None
    if text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def parse_royalty_scale_line(value):
    raw = clean_excel_text(value)
    if not raw:
        return None
    percent_match = re.search(r"(\d+(?:[.,]\d+)?)\s*%", raw)
    minimum_match = re.search(r"minimum.*?r\s*([0-9\s.,]+)", raw, re.I)
    if minimum_match:
        return {"raw": raw, "minimum": parse_money_token(minimum_match.group(1))}
    if not percent_match:
        return {"raw": raw}
    percentage = parse_money_token(percent_match.group(1)) or 0
    before_percent = raw[:percent_match.start()].strip(" -")
    money_values = [parse_money_token(item) for item in re.findall(r"R\s*[0-9][0-9\s.,]*", before_percent, flags=re.I)]
    money_values = [item for item in money_values if item is not None]
    if len(money_values) >= 2:
        amount_from, amount_to = money_values[0], money_values[1]
    elif len(money_values) == 1:
        if re.search(r"up to|tot en met", raw, re.I):
            amount_from, amount_to = 0, money_values[0]
        elif re.search(r"or more|and more|more|meer", raw, re.I):
            amount_from, amount_to = money_values[0], 999999999
        else:
            amount_from, amount_to = 0, money_values[0]
    else:
        amount_from, amount_to = 0, 999999999
    return {
        "raw": raw,
        "amount_from": amount_from,
        "amount_to": amount_to,
        "percentage": percentage,
        "label": before_percent,
    }


def find_franchise_by_name(name):
    key = normalize_franchise_key(name)
    if not key:
        return None
    franchises = Franchise.query.all()
    # exact normalized match first
    for franchise in franchises:
        if normalize_franchise_key(franchise.business_name) == key:
            return franchise
    # compare common branch aliases.  Imported contracts often contain legal
    # names such as "Martin's Funerals Alberton" or "... T/As ...", while
    # the system stores only the branch name.
    for franchise in franchises:
        aliases = [
            franchise.business_name,
            franchise.ck_business_name,
            franchise.pty_business_name,
            franchise.franchise_code,
        ]
        for alias in aliases:
            existing = normalize_franchise_key(alias)
            if key and existing and (key == existing or key in existing or existing in key):
                return franchise
    return None


def contract_group_aliases(group, worksheet):
    aliases = [group.get("name", "")]
    for r in group.get("rows", []):
        for c in (10, 12, 20, 36):
            aliases.append(clean_excel_text(worksheet.cell(r, c).value))
    cleaned = []
    seen = set()
    for alias in aliases:
        for part in re.split(r";|/|\bt/as\b|\btrading as\b", alias or "", flags=re.I):
            part = clean_franchise_name(part)
            if not part:
                continue
            key = normalize_franchise_key(part)
            if key and key not in seen:
                cleaned.append(part)
                seen.add(key)
    return cleaned


def find_franchise_for_contract_group(group, worksheet):
    for alias in contract_group_aliases(group, worksheet):
        franchise = find_franchise_by_name(alias)
        if franchise:
            return franchise, alias
    return None, ""


def ensure_franchise_owner_login(franchise):
    """Ensure a branch has one Franchise User login linked to it.

    The contract-summary import updates agreement/scale data.  It should not
    break grouped-franchise ownership, but when a branch has no franchise login
    yet we create/link the expected owner so royalties and franchise-side pages
    can find the correct user.
    """
    if not franchise:
        return None, False
    for user in getattr(franchise, "assigned_users", []) or []:
        if "Franchise User" in user_role_names(user) and not is_admin_side_user(user):
            return user, False
    return find_franchise_user_for_main_franchise(franchise)




def split_grouped_franchise_names(value):
    """Split one grouped-franchise spreadsheet cell into ordered franchise names.

    The first name in the cell is the main franchise.  Remaining names are linked
    to that main franchise user for grouped royalty calculation.
    """
    raw = clean_excel_text(value)
    if not raw:
        return []
    parts = re.split(r"[,;\n]+", raw)
    result = []
    seen = set()
    for part in parts:
        name = clean_franchise_name(part)
        if not name:
            continue
        key = normalize_franchise_key(name)
        if not key or key in seen:
            continue
        result.append(name)
        seen.add(key)
    return result


def find_franchise_user_for_main_franchise(main_franchise):
    """Find or create the franchise-side user that owns a grouped royalty set.

    Important: a franchise may already be linked to another user's group.  For
    grouped-franchise imports the FIRST name in the spreadsheet row must become
    the main owner, so do not simply return any user linked to the franchise.
    Prefer the user's own generated email/name and only accept an existing
    assigned user if that franchise is already marked primary for that user.
    """
    franchise_side_roles = {"Franchise User", "Franchise Manager", "Read Only User"}
    email = f"{slugify_email_part(main_franchise.business_name)}@martinsdirect.com"

    user = User.query.filter(db.func.lower(User.email) == email.lower()).first()
    if user and not is_admin_side_user(user):
        role = get_or_create_role("Franchise User")
        if role not in user.roles:
            user.roles.append(role)
        return user, False

    normalized_main = normalize_franchise_key(main_franchise.business_name)
    for user in getattr(main_franchise, "assigned_users", []) or []:
        if not (user_role_names(user) & franchise_side_roles) or is_admin_side_user(user):
            continue
        primary_id = db.session.execute(
            db.select(user_franchises.c.franchise_id)
            .where(user_franchises.c.user_id == user.id)
            .where(user_franchises.c.is_primary == True)
        ).scalar()
        if primary_id == main_franchise.id:
            return user, False
        # Fallback for older data where is_primary was not set but the user's
        # name/email clearly belongs to the main franchise.
        user_key = normalize_franchise_key(f"{user.name} {user.surname}")
        if normalized_main and normalized_main in user_key:
            return user, False

    display_name = re.sub(r"\s*\(F\)\s*$", "", main_franchise.business_name or "Franchise", flags=re.I).strip() or "Franchise"
    user, created = get_or_create_user(display_name, "User", email, "Franchise User", franchises=[main_franchise])
    return user, created


def remove_group_links_from_other_franchise_users(user, franchise_ids):
    """Remove grouped-branch links from other franchise-side users before re-import.

    This prevents old incorrect groups (for example Dobsonville owning Soweto)
    from remaining after the spreadsheet says Soweto is the main franchise.
    Admin/Finance users are not touched.
    """
    if not franchise_ids:
        return
    franchise_side_roles = {"Franchise User", "Franchise Manager", "Read Only User"}
    franchise_side_user_ids = [
        row[0]
        for row in db.session.query(User.id)
        .join(User.roles)
        .filter(Role.name.in_(franchise_side_roles))
        .all()
    ]
    if not franchise_side_user_ids:
        return
    db.session.execute(
        user_franchises.delete()
        .where(user_franchises.c.franchise_id.in_(franchise_ids))
        .where(user_franchises.c.user_id != user.id)
        .where(user_franchises.c.user_id.in_(franchise_side_user_ids))
    )


def set_primary_franchise_link(user, main_franchise, linked_franchises):
    """Assign linked franchises and mark the main franchise as primary in user_franchises."""
    ordered = []
    seen = set()
    for franchise in [main_franchise] + list(linked_franchises or []):
        if franchise and franchise.id not in seen:
            ordered.append(franchise)
            seen.add(franchise.id)
    remove_group_links_from_other_franchise_users(user, list(seen))
    user.assigned_franchises = ordered
    db.session.flush()
    db.session.execute(user_franchises.update().where(user_franchises.c.user_id == user.id).values(is_primary=False))
    db.session.execute(
        user_franchises.update()
        .where(user_franchises.c.user_id == user.id)
        .where(user_franchises.c.franchise_id == main_franchise.id)
        .values(is_primary=True)
    )


def clean_contact_branch_name(value):
    """Clean contact-list branch/outlet names before matching to Franchise.business_name."""
    text = normalize_contact_value(value)
    text = text.replace("*", "").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def contact_candidate_names(branch, outlet):
    """Return forgiving franchise-name candidates from contact list columns B and C.

    Column B contains the main franchise/area and column C sometimes contains the
    actual outlet.  For outlet rows such as CAPE TOWN / PAROW, the existing
    Franchise record may be stored as either 'Parow', 'Cape Town Parow', or
    'Cape Town (Parow)', so all useful variants are tried.
    """
    branch = clean_contact_branch_name(branch)
    outlet = clean_contact_branch_name(outlet)
    candidates = []
    if branch and outlet:
        # Try the outlet first because the system franchise name is commonly the
        # outlet name, then fall back to combined names and the main branch.
        candidates.extend([
            outlet,
            f"{branch} {outlet}",
            f"{branch} - {outlet}",
            f"{branch} ({outlet})",
            branch,
        ])
    elif outlet:
        candidates.append(outlet)
    elif branch:
        candidates.append(branch)
    # Remove duplicates while preserving order.
    seen = set()
    result = []
    for item in candidates:
        key = normalize_franchise_key(item)
        if key and key not in seen:
            result.append(item)
            seen.add(key)
    return result


def find_franchise_by_candidates(candidates):
    """Find a franchise using exact, contains, and safe fuzzy matching.

    Contact-list branch names do not always match the saved franchise name exactly.
    Examples: '*CAPE TOWN' + 'PAROW' may be saved as 'Cape Town Parow',
    and some saved names include 'User', '(F)', 'Martin's Funerals', etc.
    """
    cleaned_candidates = []
    seen = set()
    for candidate in candidates:
        key = normalize_franchise_key(candidate)
        if key and key not in seen:
            cleaned_candidates.append((candidate, key))
            seen.add(key)

    franchises = Franchise.query.all()
    franchise_keys = [(franchise, normalize_franchise_key(franchise.business_name)) for franchise in franchises]

    # 1) Exact normalized match.
    for candidate, key in cleaned_candidates:
        for franchise, existing_key in franchise_keys:
            if key == existing_key:
                return franchise, candidate

    # 2) Containment match. Prefer longer keys to avoid broad matches.
    for candidate, key in sorted(cleaned_candidates, key=lambda item: len(item[1]), reverse=True):
        for franchise, existing_key in franchise_keys:
            if key and existing_key and (key in existing_key or existing_key in key):
                return franchise, candidate

    # 3) Safe fuzzy match for spelling/spacing variations.
    best = (None, "", 0.0)
    for candidate, key in cleaned_candidates:
        for franchise, existing_key in franchise_keys:
            if not key or not existing_key:
                continue
            ratio = SequenceMatcher(None, key, existing_key).ratio()
            if ratio > best[2]:
                best = (franchise, candidate, ratio)
    if best[0] is not None and best[2] >= 0.86:
        return best[0], f"{best[1]} (fuzzy {best[2]:.0%})"

    return None, ""




def clean_shareholder_name(value):
    """Return a clean shareholder name from column N.

    The contract sheet often stores shareholders as numbered values such as
    '1. Jan van Wyk'. Only the primary/first shareholder is imported into the
    single Franchisee Name/Surname fields.
    """
    text = clean_excel_text(value)
    if not text:
        return ""
    text = text.replace("\r", "\n").replace(";", "\n")
    candidates = []
    for part in re.split(r"\n+|(?=\b\d+\s*[.)]\s+)", text):
        part = clean_excel_text(part)
        if part:
            candidates.append(part)
    if not candidates:
        candidates = [text]
    # Prefer the explicitly numbered primary shareholder, otherwise use the first non-empty value.
    chosen = ""
    for candidate in candidates:
        if re.match(r"^1\s*[.)]\s+", candidate):
            chosen = candidate
            break
    if not chosen:
        chosen = candidates[0]
    chosen = re.sub(r"^\s*\d+\s*[.)]\s*", "", chosen).strip()
    return clean_excel_text(chosen)


def first_shareholder_name(values):
    for value in values:
        name = clean_shareholder_name(value)
        if name:
            return name
    return ""

def split_person_name(full_name):
    text = clean_excel_text(full_name)
    # The contact sheet sometimes includes a 24H number inside NAME & SURNAME.
    text = re.sub(r"24\s*h.*$", "", text, flags=re.I).strip()
    if not text:
        return "", ""
    parts = text.split()
    if len(parts) == 1:
        return parts[0].title(), ""
    return " ".join(parts[:-1]).title(), parts[-1].title()


def normalize_contact_value(value):
    text = clean_excel_text(value)
    text = text.replace(" ", " ")
    return " ".join(text.split())

def build_contract_summary_groups(worksheet):
    """Build current contract-summary groups from B2:B375.

    The contract workbook stores royalty-scale continuation rows with a blank
    franchise name in column B.  A franchise can also appear more than once when
    an old agreement and a newer agreement are both in the file.  In that case
    the import must keep the newest agreement block and replace the older data,
    otherwise the database will not reflect the latest Excel upload.
    """
    blocks = []
    current_block = None
    last_row = min(worksheet.max_row, 375)

    for row_number in range(2, last_row + 1):
        franchise_name = clean_excel_text(worksheet.cell(row_number, 2).value)
        if franchise_name:
            key = normalize_franchise_key(franchise_name)
            if not key or key in {"total", "totals", "data"}:
                current_block = None
                continue
            current_block = {"name": franchise_name, "key": key, "rows": []}
            blocks.append(current_block)

        if current_block:
            row_values = [worksheet.cell(row_number, c).value for c in (4, 5, 10, 11, 12, 13, 14, 18)]
            if franchise_name or any(clean_excel_text(v) for v in row_values):
                current_block["rows"].append(row_number)

    selected_by_key = {}
    selected_order = []
    for block in blocks:
        rows = block["rows"]
        block_start_date = newest_contract_date(worksheet.cell(r, 4).value for r in rows)
        existing = selected_by_key.get(block["key"])
        if not existing:
            block["start_date_for_selection"] = block_start_date
            selected_by_key[block["key"]] = block
            selected_order.append(block["key"])
            continue

        existing_date = existing.get("start_date_for_selection")
        should_replace = False
        if block_start_date and existing_date:
            should_replace = block_start_date >= existing_date
        elif block_start_date and not existing_date:
            should_replace = True

        if should_replace:
            block["start_date_for_selection"] = block_start_date
            selected_by_key[block["key"]] = block

    return [selected_by_key[key] for key in selected_order if key in selected_by_key]



@admin_bp.route("/imports")
@login_required
def imports_data():
    if not can_view_imports_data():
        abort(403)
    role_names = current_user_role_names()
    return render_template(
        "admin/imports_data.html",
        is_import_admin="Admin" in role_names,
        is_import_finance=bool(role_names & {"Finance Manager", "Finance Assistant"}),
    )




def _import_job_report(job):
    import json
    if not job or not getattr(job, "extra_json", None):
        return {}
    try:
        return json.loads(job.extra_json or "{}")
    except Exception:
        return {"raw": job.extra_json}


@admin_bp.route("/imports/centre")
@login_required
def import_centre():
    if not can_view_imports_data():
        abort(403)
    jobs = ImportJob.query.order_by(ImportJob.started_at.desc()).limit(50).all()
    decorated_jobs = []
    for job in jobs:
        report = _import_job_report(job)
        decorated_jobs.append({"job": job, "report": report})
    return render_template("admin/import_centre.html", jobs=decorated_jobs)


@admin_bp.route("/imports/centre/<int:job_id>")
@login_required
def import_centre_detail(job_id):
    if not can_view_imports_data():
        abort(403)
    job = ImportJob.query.get_or_404(job_id)
    report = _import_job_report(job)
    return render_template("admin/import_centre_detail.html", job=job, report=report)


@admin_bp.route("/imports/grouped-franchises", methods=["GET", "POST"])
@login_required
def import_grouped_franchises():
    if not is_current_user_admin():
        abort(403)
    if request.method == "POST":
        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Please upload the grouped franchises Excel file.", "danger")
            return redirect(url_for("admin.import_grouped_franchises"))
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            flash(f"Could not read Excel file: {exc}", "danger")
            return redirect(url_for("admin.import_grouped_franchises"))

        processed = matched = created_users = 0
        grouped = []
        unmatched = []

        # Read the workbook in the exact format supplied by the user:
        #   Column A = main franchise user / main franchise
        #   Columns B:E = franchise users/franchises to link under column A
        #   Rows 2:14 contain the grouping data; row 1 is a heading row.
        #
        # The important rule is that Column A always remains the main franchise.
        # Columns B:E are linked to Column A, even if the same franchise used to
        # be linked somewhere else from a previous import.
        import_rows = []
        all_matched_franchise_ids = set()
        max_row = min(worksheet.max_row, 14)
        max_col = min(max(worksheet.max_column, 5), 5)
        for row_number in range(2, max_row + 1):
            main_name = clean_franchise_name(worksheet.cell(row_number, 1).value)
            if not main_name:
                continue

            linked_names = []
            seen_names = {normalize_franchise_key(main_name)}
            for col in range(2, max_col + 1):
                cell_value = worksheet.cell(row_number, col).value
                # A linked cell may contain one name or comma/semi-colon separated
                # names; split it safely while preserving the main in column A.
                for candidate in split_grouped_franchise_names(cell_value):
                    key = normalize_franchise_key(candidate)
                    if key and key not in seen_names:
                        linked_names.append(candidate)
                        seen_names.add(key)

            names = [main_name] + linked_names
            processed += 1
            resolved = []
            row_unmatched = []

            main_franchise = find_franchise_by_name(main_name)
            if not main_franchise:
                row_unmatched.append({
                    "row": row_number,
                    "name": main_name,
                    "reason": "Main franchise not found",
                    "linked": names,
                })
                unmatched.extend(row_unmatched)
                continue

            resolved.append(main_franchise)
            all_matched_franchise_ids.add(main_franchise.id)

            for name in linked_names:
                franchise = find_franchise_by_name(name)
                if franchise:
                    if franchise.id not in {item.id for item in resolved}:
                        resolved.append(franchise)
                        all_matched_franchise_ids.add(franchise.id)
                else:
                    row_unmatched.append({
                        "row": row_number,
                        "name": name,
                        "reason": "Linked franchise not found",
                        "linked": names,
                    })

            import_rows.append({
                "row": row_number,
                "names": names,
                "main": main_franchise,
                "franchises": resolved,
                "unmatched": row_unmatched,
            })
            unmatched.extend(row_unmatched)

        # Remove existing franchise-side links for every branch mentioned in the
        # sheet before adding the new links.  This is the important fix: old
        # groups such as Dobsonville -> Soweto cannot remain when the Excel row
        # says Soweto must be the main branch.  Admin/Finance users are not
        # touched.
        if all_matched_franchise_ids:
            franchise_side_roles = {"Franchise User", "Franchise Manager", "Read Only User"}
            franchise_side_user_ids = [
                row[0]
                for row in db.session.query(User.id)
                .join(User.roles)
                .filter(Role.name.in_(franchise_side_roles))
                .all()
            ]
            if franchise_side_user_ids:
                db.session.execute(
                    user_franchises.delete()
                    .where(user_franchises.c.franchise_id.in_(list(all_matched_franchise_ids)))
                    .where(user_franchises.c.user_id.in_(franchise_side_user_ids))
                )
                db.session.flush()

        for item in import_rows:
            main_franchise = item["main"]
            linked_franchises = item["franchises"]
            user, created = find_franchise_user_for_main_franchise(main_franchise)
            if created:
                created_users += 1
            set_primary_franchise_link(user, main_franchise, linked_franchises)
            matched += 1
            grouped.append({
                "row": item["row"],
                "main": main_franchise.business_name,
                "user": user.full_name,
                "email": user.email,
                "linked": [branch.business_name for branch in linked_franchises],
                "unmatched": [entry["name"] for entry in item["unmatched"]],
            })

        db.session.commit()
        log_action("Imports & Data", "Imported grouped franchises", f"Processed: {processed}, matched groups: {matched}, unmatched: {len(unmatched)}")
        flash(f"Grouped franchises imported. {matched} group(s) updated.", "success")
        return render_template(
            "admin/import_grouped_franchises.html",
            import_complete=True,
            processed=processed,
            matched=matched,
            created_users=created_users,
            grouped=grouped,
            unmatched=unmatched,
        )

    return render_template(
        "admin/import_grouped_franchises.html",
        import_complete=False,
        processed=0,
        matched=0,
        created_users=0,
        grouped=[],
        unmatched=[],
    )


@admin_bp.route("/imports/contract-summary", methods=["GET", "POST"])
@login_required
def import_contract_summary():
    if not is_current_user_admin():
        abort(403)
    if request.method == "POST":
        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Please upload the contract summary Excel file.", "danger")
            return redirect(url_for("admin.import_contract_summary"))
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            flash(f"Could not read Excel file: {exc}", "danger")
            return redirect(url_for("admin.import_contract_summary"))

        groups = build_contract_summary_groups(worksheet)
        from app.import_progress import start_import_job, update_import_job
        job = start_import_job("contract_summary", uploaded_file.filename, total_steps=max(len(groups), 1))
        processed = matched = updated_scales = 0
        unmatched = []
        created_or_linked_users = 0
        for group in groups:
            processed += 1
            franchise, matched_alias = find_franchise_for_contract_group(group, worksheet)
            update_import_job(job, processed, f"Processing {group['name']} ({processed}/{len(groups)})", commit=True)
            if not franchise:
                unmatched.append({"name": group["name"], "aliases": contract_group_aliases(group, worksheet)})
                continue
            matched += 1
            rows = group["rows"]

            # Full refresh from the latest uploaded Excel file.
            # Dates are overwritten so changed agreement dates in the workbook update the Franchise Details page.
            start_date = newest_contract_date(worksheet.cell(r, 4).value for r in rows)
            end_date = newest_contract_date(worksheet.cell(r, 5).value for r in rows)
            franchise.agreement_start_date = start_date
            franchise.agreement_end_date = end_date
            set_auto_gross_method_from_agreement(franchise)

            ck_business_name = unique_join(worksheet.cell(r, 10).value for r in rows)
            ck_number = unique_join(worksheet.cell(r, 11).value for r in rows)
            pty_business_name = unique_join(worksheet.cell(r, 12).value for r in rows)
            pty_number = unique_join(worksheet.cell(r, 13).value for r in rows)
            franchisee = first_shareholder_name(worksheet.cell(r, 14).value for r in rows)

            # Full sync for contract master data: if the new file changed a value, replace it.
            franchise.ck_business_name = ck_business_name
            franchise.ck_number = ck_number
            franchise.pty_business_name = pty_business_name
            franchise.pty_number = pty_number
            if franchisee:
                first_name, surname = split_person_name(franchisee)
                franchise.franchisee_name = first_name
                franchise.franchisee_surname = surname
            else:
                franchise.franchisee_name = ""
                franchise.franchisee_surname = ""

            raw_scale_lines = []
            parsed_rows = []
            minimum = None
            for r in rows:
                parsed = parse_royalty_scale_line(worksheet.cell(r, 18).value)
                if not parsed:
                    continue
                raw = parsed.get("raw", "")
                if raw:
                    raw_scale_lines.append(raw)
                if parsed.get("minimum") is not None:
                    minimum = parsed["minimum"]
                    continue
                if "percentage" in parsed:
                    parsed_rows.append(parsed)

            # Full database sync from the latest uploaded Excel file.
            # This clears old scale rows first, then rebuilds them exactly from the new file.
            updated_scale_rows = sync_royalty_scales_from_contract_file(franchise, parsed_rows, raw_scale_lines, minimum)
            if updated_scale_rows:
                updated_scales += 1
            owner, owner_created = ensure_franchise_owner_login(franchise)
            if owner_created:
                created_or_linked_users += 1

        # Recalculate existing monthly figures because the agreement date controls
        # whether the franchise uses Gross = New Gross Method or Gross = Old.
        from app.monthly.routes import recalculate_monthly_figure
        for franchise in Franchise.query.all():
            for figure in MonthlyFigure.query.filter_by(franchise_id=franchise.id).all():
                recalculate_monthly_figure(figure)

        db.session.commit()
        update_import_job(job, job.total_steps, f"Import complete. {matched} franchises matched; {updated_scales} royalty scales updated.", status="completed", commit=True)
        log_action("Imports & Data", "Imported contract summary", f"Processed: {processed}, matched: {matched}, unmatched: {len(unmatched)}")
        flash(f"Contract summary import complete. {processed} franchises processed, {matched} matched, {len(unmatched)} unmatched, {updated_scales} royalty scales updated.", "success")
        return render_template(
            "admin/import_contract_summary.html",
            import_complete=True,
            processed=processed,
            matched=matched,
            unmatched=unmatched,
            updated_scales=updated_scales,
            created_or_linked_users=created_or_linked_users,
        )

    return render_template("admin/import_contract_summary.html", import_complete=False, processed=0, matched=0, unmatched=[], updated_scales=0, created_or_linked_users=0)


@admin_bp.route("/imports/status/latest")
@login_required
def latest_import_status():
    if not can_view_imports_data():
        abort(403)
    job = ImportJob.query.order_by(ImportJob.started_at.desc()).first()
    if not job:
        return jsonify({"status": "none", "progress_percent": 0, "message": "No imports started yet."})
    return jsonify({
        "id": job.id,
        "kind": job.kind,
        "filename": job.filename,
        "status": job.status,
        "message": job.message,
        "current_step": job.current_step,
        "total_steps": job.total_steps,
        "progress_percent": job.progress_percent,
        "started_at": job.started_at.isoformat() if job.started_at else None,
        "finished_at": job.finished_at.isoformat() if job.finished_at else None,
    })





def apply_contact_data_to_franchise(franchise, contact_name="", office_number="", email="", address=""):
    """Apply imported/manual contact-list values to one Franchise record."""
    contact_name = normalize_contact_value(contact_name)
    office_number = normalize_contact_value(office_number)
    email = normalize_contact_value(email)
    address = normalize_contact_value(address)

    # Full sync: every new upload overwrites the old database values.
    franchise.office_number = office_number
    franchise.after_hours_number = office_number
    franchise.franchisee_cell = office_number
    franchise.franchisee_email = email
    franchise.public_email = email
    franchise.office_address = address
    first_name, surname = split_person_name(contact_name)
    franchise.franchisee_name = first_name or ""
    franchise.franchisee_surname = surname or ""

@admin_bp.route("/imports/contact-list", methods=["GET", "POST"])
@login_required
def import_contact_list():
    if not is_current_user_admin():
        abort(403)
    if request.method == "POST":
        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Please upload the Martins Funerals contact list Excel file.", "danger")
            return redirect(url_for("admin.import_contact_list"))
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            flash(f"Could not read Excel file: {exc}", "danger")
            return redirect(url_for("admin.import_contact_list"))

        processed = matched = 0
        unmatched = []
        updated = []
        # The uploaded contact list currently has headers in row 1 and data from row 2.
        # Process all populated rows, not only the first 76, because some files include
        # more than 76 outlets/branches.
        last_row = worksheet.max_row
        for row_number in range(2, last_row + 1):
            branch = normalize_contact_value(worksheet.cell(row_number, 2).value)
            outlet = normalize_contact_value(worksheet.cell(row_number, 3).value)
            if not branch and not outlet:
                continue
            processed += 1
            candidates = contact_candidate_names(branch, outlet)
            franchise, matched_name = find_franchise_by_candidates(candidates)
            display_name = " / ".join([item for item in [branch, outlet] if item])
            if not franchise:
                unmatched.append({
                    "row": row_number,
                    "name": display_name or f"Row {row_number}",
                    "branch": branch,
                    "outlet": outlet,
                    "contact_name": normalize_contact_value(worksheet.cell(row_number, 4).value),
                    "office_number": normalize_contact_value(worksheet.cell(row_number, 6).value),
                    "email": normalize_contact_value(worksheet.cell(row_number, 7).value),
                    "address": normalize_contact_value(worksheet.cell(row_number, 8).value),
                    "tried": ", ".join(candidates),
                })
                continue

            matched += 1
            contact_name = normalize_contact_value(worksheet.cell(row_number, 4).value)
            # User-approved mapping for this file:
            # B/C = franchise name/outlet, F = office number, 24-hour number and cell number,
            # G = franchisee/public email, H = office address.
            office_number = normalize_contact_value(worksheet.cell(row_number, 6).value)
            email = normalize_contact_value(worksheet.cell(row_number, 7).value)
            address = normalize_contact_value(worksheet.cell(row_number, 8).value)

            # Store the uploaded contact information on the Franchise record so every
            # linked franchise user sees it on the Franchise Details page.
            apply_contact_data_to_franchise(franchise, contact_name, office_number, email, address)

            updated.append({
                "row": row_number,
                "spreadsheet_name": display_name,
                "matched_franchise": franchise.business_name,
                "matched_by": matched_name,
                "email": email,
                "office_number": office_number,
            })

        db.session.commit()
        log_action("Imports & Data", "Imported contact list", f"Processed: {processed}, matched: {matched}, unmatched: {len(unmatched)}")
        flash(f"Contact list import complete. {processed} rows processed, {matched} matched, {len(unmatched)} unmatched.", "success")
        return render_template(
            "admin/import_contact_list.html",
            import_complete=True,
            processed=processed,
            matched=matched,
            unmatched=unmatched,
            updated=updated,
        )

    return render_template("admin/import_contact_list.html", import_complete=False, processed=0, matched=0, unmatched=[], updated=[])



@admin_bp.route("/imports/contact-list/manual", methods=["GET", "POST"])
@login_required
def manual_allocate_contact_list_row():
    if not is_current_user_admin():
        abort(403)
    franchises = Franchise.query.order_by(Franchise.business_name).all()
    if request.method == "POST":
        franchise_id = request.form.get("franchise_id", type=int)
        franchise = Franchise.query.get(franchise_id) if franchise_id else None
        if not franchise:
            flash("Please select the correct franchise before saving.", "danger")
            return redirect(url_for("admin.import_contact_list"))

        contact_name = request.form.get("contact_name", "")
        office_number = request.form.get("office_number", "")
        email = request.form.get("email", "")
        address = request.form.get("address", "")
        spreadsheet_name = request.form.get("spreadsheet_name", "")
        row_number = request.form.get("row_number", "")

        apply_contact_data_to_franchise(franchise, contact_name, office_number, email, address)
        db.session.commit()
        log_action(
            "Imports & Data",
            "Manually allocated contact-list row",
            f"Row: {row_number}; Spreadsheet: {spreadsheet_name}; Franchise: {franchise.business_name}",
        )
        flash(f"Contact-list row allocated to {franchise.business_name} and saved to Franchise Details.", "success")
        return redirect(url_for("admin.import_contact_list"))

    row_data = {
        "row_number": request.args.get("row", ""),
        "spreadsheet_name": request.args.get("name", ""),
        "branch": request.args.get("branch", ""),
        "outlet": request.args.get("outlet", ""),
        "contact_name": request.args.get("contact_name", ""),
        "office_number": request.args.get("office_number", ""),
        "email": request.args.get("email", ""),
        "address": request.args.get("address", ""),
        "tried": request.args.get("tried", ""),
    }
    return render_template("admin/manual_allocate_contact.html", franchises=franchises, row_data=row_data)

@admin_bp.route("/roles")
@login_required
@permission_required("user_roles:view")
def roles():
    roles = Role.query.order_by(Role.name).all()
    return render_template("admin/roles.html", roles=roles)


@admin_bp.route("/roles/<int:role_id>", methods=["GET", "POST"])
@login_required
@permission_required("user_roles:edit")
def edit_role(role_id):
    role = Role.query.get_or_404(role_id)
    permissions = Permission.query.order_by(Permission.sort_order).all()
    grouped = defaultdict(list)
    for permission in permissions:
        grouped[permission.module].append(permission)

    if request.method == "POST":
        selected_ids = [int(item) for item in request.form.getlist("permission_ids")]
        role.permissions = Permission.query.filter(Permission.id.in_(selected_ids)).all() if selected_ids else []
        log_action("User Roles", "Updated role permissions", f"Role: {role.name}")
        db.session.commit()
        flash(f"Permissions updated for {role.name}.", "success")
        return redirect(url_for("admin.roles"))

    selected = {permission.id for permission in role.permissions}
    return render_template("admin/edit_role.html", role=role, grouped_permissions=grouped, actions=ACTIONS, selected=selected)





def can_view_operations_centre():
    names = current_user_role_names()
    return bool(names & {"Admin", "Super Admin", "Finance Manager"}) or current_user.has_permission("system_administration:view")


def _ops_count(table_name):
    try:
        return int(db.session.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar() or 0)
    except Exception:
        return 0


def _latest_import_period():
    row = db.session.execute(text("""
        SELECT year, month, COUNT(*) AS rows
        FROM monthly_figures
        GROUP BY year, month
        ORDER BY year DESC, month DESC
        LIMIT 1
    """)).mappings().first()
    if not row:
        now = datetime.utcnow()
        return {"year": now.year, "month": now.month, "rows": 0}
    return dict(row)



def _safe_scalar(sql, params=None, default=0):
    """Run a small aggregate safely for executive dashboards."""
    try:
        value = db.session.execute(text(sql), params or {}).scalar()
        return value if value is not None else default
    except Exception:
        db.session.rollback()
        return default


def _safe_rows(sql, params=None):
    try:
        return [dict(row) for row in db.session.execute(text(sql), params or {}).mappings().all()]
    except Exception:
        db.session.rollback()
        return []


def _exec_latest_period():
    row = _safe_rows("""
        SELECT year, month, COUNT(*) AS rows
        FROM monthly_figures
        GROUP BY year, month
        ORDER BY year DESC, month DESC
        LIMIT 1
    """)
    if row:
        return row[0]
    now = datetime.utcnow()
    return {"year": now.year, "month": now.month, "rows": 0}


def _exec_previous_period(year, month):
    if not year or not month:
        return {"year": None, "month": None}
    if int(month) == 1:
        return {"year": int(year) - 1, "month": 12}
    return {"year": int(year), "month": int(month) - 1}


def _period_totals(year, month):
    rows = _safe_rows("""
        SELECT
            COUNT(*) AS rows,
            COUNT(DISTINCT franchise_id) AS franchise_count,
            COALESCE(SUM(gross_turnover), 0) AS gross_turnover,
            COALESCE(SUM(royalty_amount), 0) AS royalty_amount,
            COALESCE(SUM(payover), 0) AS payover,
            COALESCE(SUM(sales), 0) AS sales,
            COALESCE(SUM(number_of_funerals), 0) AS funerals,
            COALESCE(SUM(insurance_joinings), 0) AS joinings,
            COALESCE(AVG(NULLIF(royalty_percentage, 0)), 0) AS avg_royalty_percentage
        FROM monthly_figures
        WHERE year = :year AND month = :month
    """, {"year": year, "month": month})
    return rows[0] if rows else {}


def _delta(current, previous):
    try:
        current = float(current or 0)
        previous = float(previous or 0)
    except Exception:
        return {"amount": 0, "percent": 0, "tone": "neutral"}
    amount = current - previous
    percent = (amount / previous * 100) if previous else (100 if current else 0)
    return {"amount": amount, "percent": percent, "tone": "ok" if amount >= 0 else "danger"}


@admin_bp.route("/executive-dashboard")
@login_required
def executive_dashboard():
    """Phase 10 executive dashboard: high-level KPIs and operational health."""
    if not can_view_operations_centre():
        abort(403)

    latest_period = _exec_latest_period()
    selected_year = int(request.args.get("year") or latest_period.get("year") or datetime.utcnow().year)
    selected_month = int(request.args.get("month") or latest_period.get("month") or datetime.utcnow().month)
    previous_period = _exec_previous_period(selected_year, selected_month)

    totals = _period_totals(selected_year, selected_month)
    previous_totals = _period_totals(previous_period.get("year"), previous_period.get("month")) if previous_period.get("year") else {}

    kpis = [
        {"label": "Gross Turnover", "value": totals.get("gross_turnover", 0), "format": "money", "delta": _delta(totals.get("gross_turnover"), previous_totals.get("gross_turnover"))},
        {"label": "Royalties", "value": totals.get("royalty_amount", 0), "format": "money", "delta": _delta(totals.get("royalty_amount"), previous_totals.get("royalty_amount"))},
        {"label": "Payover", "value": totals.get("payover", 0), "format": "money", "delta": _delta(totals.get("payover"), previous_totals.get("payover"))},
        {"label": "Franchises Reporting", "value": totals.get("franchise_count", 0), "format": "number", "delta": _delta(totals.get("franchise_count"), previous_totals.get("franchise_count"))},
        {"label": "Funerals", "value": totals.get("funerals", 0), "format": "number", "delta": _delta(totals.get("funerals"), previous_totals.get("funerals"))},
        {"label": "Joinings", "value": totals.get("joinings", 0), "format": "number", "delta": _delta(totals.get("joinings"), previous_totals.get("joinings"))},
    ]

    top_franchises = _safe_rows("""
        SELECT f.id, f.business_name, mf.gross_turnover, mf.royalty_amount, mf.payover,
               mf.number_of_funerals, mf.insurance_joinings
        FROM monthly_figures mf
        JOIN franchises f ON f.id = mf.franchise_id
        WHERE mf.year = :year AND mf.month = :month
        ORDER BY COALESCE(mf.gross_turnover, 0) DESC, f.business_name
        LIMIT 10
    """, {"year": selected_year, "month": selected_month})

    bottom_franchises = _safe_rows("""
        SELECT f.id, f.business_name, mf.gross_turnover, mf.royalty_amount, mf.payover,
               mf.number_of_funerals, mf.insurance_joinings
        FROM monthly_figures mf
        JOIN franchises f ON f.id = mf.franchise_id
        WHERE mf.year = :year AND mf.month = :month
        ORDER BY COALESCE(mf.gross_turnover, 0) ASC, f.business_name
        LIMIT 10
    """, {"year": selected_year, "month": selected_month})

    province_rows = _safe_rows("""
        WITH province_map AS (
            SELECT franchise_id, MAX(NULLIF(province, '')) AS province
            FROM heatmap_records
            GROUP BY franchise_id
        )
        SELECT COALESCE(NULLIF(f.province, ''), pm.province, 'Unassigned') AS province,
               COUNT(DISTINCT mf.franchise_id) AS franchises,
               COALESCE(SUM(mf.gross_turnover), 0) AS gross_turnover,
               COALESCE(SUM(mf.royalty_amount), 0) AS royalty_amount,
               COALESCE(SUM(mf.number_of_funerals), 0) AS funerals,
               COALESCE(SUM(mf.insurance_joinings), 0) AS joinings
        FROM monthly_figures mf
        JOIN franchises f ON f.id = mf.franchise_id
        LEFT JOIN province_map pm ON pm.franchise_id = mf.franchise_id
        WHERE mf.year = :year AND mf.month = :month
        GROUP BY COALESCE(NULLIF(f.province, ''), pm.province, 'Unassigned')
        ORDER BY gross_turnover DESC
        LIMIT 12
    """, {"year": selected_year, "month": selected_month})

    periods = _safe_rows("""
        SELECT year, month, COUNT(*) AS rows,
               COALESCE(SUM(gross_turnover), 0) AS gross_turnover,
               COALESCE(SUM(royalty_amount), 0) AS royalty_amount
        FROM monthly_figures
        GROUP BY year, month
        ORDER BY year DESC, month DESC
        LIMIT 12
    """)

    latest_imports = ImportJob.query.order_by(ImportJob.started_at.desc()).limit(8).all()
    running_imports = ImportJob.query.filter(ImportJob.status.in_(["queued", "running", "processing", "validating", "publishing"])).order_by(ImportJob.started_at.desc()).limit(8).all()
    needs_review_imports = ImportJob.query.filter(ImportJob.status.in_(["needs_review", "failed", "warning"])).order_by(ImportJob.started_at.desc()).limit(8).all()
    worker_heartbeats = WorkerHeartbeat.query.order_by(WorkerHeartbeat.heartbeat_at.desc()).limit(5).all()
    online_workers = [w for w in worker_heartbeats if w.is_online and w.status != "stopped"]

    try:
        cache = cache_stats()
    except Exception:
        cache = {"valid": 0, "invalidated": 0}
    bi_summary = {}
    try:
        from app.business_intelligence import get_intelligence_summary
        bi_summary = get_intelligence_summary(selected_year, selected_month)
    except Exception:
        db.session.rollback()
        bi_summary = {"avg_score": 0, "status": {}, "insights": []}
    event_stats_data = {}
    try:
        from app.events import event_stats
        event_stats_data = event_stats()
    except Exception:
        event_stats_data = {}

    warnings = [
        {"label": "Imports needing review", "value": len(needs_review_imports), "tone": "danger" if needs_review_imports else "ok", "url": url_for("admin.import_centre")},
        {"label": "Running import jobs", "value": len(running_imports), "tone": "warning" if running_imports else "ok", "url": url_for("admin.database_diagnostics")},
        {"label": "Royalty rows needing review", "value": _safe_scalar("SELECT COUNT(*) FROM royalty_calculation_snapshots WHERE status = 'needs_review'", default=0), "tone": "danger", "url": url_for("admin.royalty_management")},
        {"label": "Rows with gross turnover but zero royalty", "value": _safe_scalar("SELECT COUNT(*) FROM monthly_figures WHERE COALESCE(gross_turnover,0) > 0 AND COALESCE(royalty_amount,0) = 0", default=0), "tone": "warning", "url": url_for("admin.database_diagnostics")},
        {"label": "Missing agreement dates", "value": _safe_scalar("SELECT COUNT(*) FROM franchises WHERE agreement_start_date IS NULL OR agreement_end_date IS NULL", default=0), "tone": "warning", "url": url_for("admin.database_diagnostics")},
        {"label": "Failed events", "value": event_stats_data.get("failed", 0), "tone": "danger", "url": url_for("admin.database_diagnostics")},
        {"label": "Online workers", "value": len(online_workers), "tone": "ok" if online_workers else "warning", "url": url_for("admin.database_diagnostics")},
        {"label": "Valid cache rows", "value": cache.get("valid", 0), "tone": "ok" if cache.get("valid", 0) else "warning", "url": url_for("admin.database_diagnostics")},
        {"label": "BI health score", "value": f"{bi_summary.get('avg_score', 0):.1f}%", "tone": "ok" if bi_summary.get("avg_score", 0) >= 70 else "warning", "url": url_for("admin.business_intelligence", year=selected_year, month=selected_month)},
    ]

    quick_links = [
        {"label": "Import Centre", "detail": "Review uploads and import results", "url": url_for("admin.import_centre")},
        {"label": "Royalty Management", "detail": "Recalculate and audit royalties", "url": url_for("admin.royalty_management")},
        {"label": "Business Intelligence", "detail": "Health scoring, trends and insights", "url": url_for("admin.business_intelligence", year=selected_year, month=selected_month)},
        {"label": "Insight Explanations", "detail": "Plain-language explanations and monthly summaries", "url": url_for("admin.insights_dashboard", year=selected_year, month=selected_month)},
        {"label": "Performance Graphs", "detail": "View company and franchise graphs", "url": url_for("performance.graphs")},
        {"label": "Leaderboard", "detail": "Company-wide franchise ranking", "url": url_for("performance.index")},
        {"label": "Database Diagnostics", "detail": "Find missing links and data issues", "url": url_for("admin.database_diagnostics")},
    ]

    return render_template(
        "admin/executive_dashboard.html",
        selected_year=selected_year,
        selected_month=selected_month,
        latest_period=latest_period,
        previous_period=previous_period,
        totals=totals,
        previous_totals=previous_totals,
        kpis=kpis,
        top_franchises=top_franchises,
        bottom_franchises=bottom_franchises,
        province_rows=province_rows,
        periods=periods,
        latest_imports=latest_imports,
        running_imports=running_imports,
        needs_review_imports=needs_review_imports,
        worker_heartbeats=worker_heartbeats,
        online_workers=online_workers,
        cache=cache,
        event_stats=event_stats_data,
        bi_summary=bi_summary,
        warnings=warnings,
        quick_links=quick_links,
    )


@admin_bp.route("/business-intelligence")
@login_required
def business_intelligence():
    """Phase 11 Enterprise Business Intelligence dashboard."""
    if not can_view_operations_centre():
        abort(403)

    from app.business_intelligence import latest_period, get_intelligence_summary

    latest = latest_period()
    selected_year = int(request.args.get("year") or latest.get("year") or datetime.utcnow().year)
    selected_month = int(request.args.get("month") or latest.get("month") or datetime.utcnow().month)
    summary = get_intelligence_summary(selected_year, selected_month)

    health_rows = FranchiseHealthSnapshot.query.filter_by(year=selected_year, month=selected_month)\
        .join(Franchise)\
        .order_by(FranchiseHealthSnapshot.health_score.asc(), Franchise.business_name.asc())\
        .limit(40).all()

    top_growth = FranchiseHealthSnapshot.query.filter_by(year=selected_year, month=selected_month)\
        .join(Franchise)\
        .order_by(FranchiseHealthSnapshot.growth_percent.desc(), Franchise.business_name.asc())\
        .limit(10).all()

    biggest_decline = FranchiseHealthSnapshot.query.filter_by(year=selected_year, month=selected_month)\
        .join(Franchise)\
        .order_by(FranchiseHealthSnapshot.growth_percent.asc(), Franchise.business_name.asc())\
        .limit(10).all()

    target_misses = FranchiseHealthSnapshot.query.filter_by(year=selected_year, month=selected_month)\
        .filter(FranchiseHealthSnapshot.target_amount > 0)\
        .filter(FranchiseHealthSnapshot.target_achievement_percent < 80)\
        .join(Franchise)\
        .order_by(FranchiseHealthSnapshot.target_achievement_percent.asc(), Franchise.business_name.asc())\
        .limit(15).all()

    province_rows = _safe_rows("""
        WITH province_map AS (
            SELECT franchise_id, MAX(NULLIF(province, '')) AS province
            FROM heatmap_records
            GROUP BY franchise_id
        )
        SELECT COALESCE(NULLIF(f.province, ''), pm.province, 'Unassigned') AS province,
               COUNT(*) AS franchises,
               COALESCE(AVG(fhs.health_score), 0) AS avg_health_score,
               COALESCE(SUM(fhs.gross_turnover), 0) AS gross_turnover,
               COALESCE(SUM(fhs.royalty_amount), 0) AS royalty_amount,
               COALESCE(AVG(fhs.growth_percent), 0) AS avg_growth_percent,
               SUM(CASE WHEN fhs.health_status = 'critical' THEN 1 ELSE 0 END) AS critical_count,
               SUM(CASE WHEN fhs.health_status = 'watch' THEN 1 ELSE 0 END) AS watch_count,
               SUM(CASE WHEN fhs.health_status = 'healthy' THEN 1 ELSE 0 END) AS healthy_count
        FROM franchise_health_snapshots fhs
        JOIN franchises f ON f.id = fhs.franchise_id
        LEFT JOIN province_map pm ON pm.franchise_id = fhs.franchise_id
        WHERE fhs.year = :year AND fhs.month = :month
        GROUP BY COALESCE(NULLIF(f.province, ''), pm.province, 'Unassigned')
        ORDER BY avg_health_score ASC, gross_turnover DESC
    """, {"year": selected_year, "month": selected_month})

    periods = _safe_rows("""
        SELECT year, month, COUNT(*) AS snapshots, COALESCE(AVG(health_score), 0) AS avg_health_score
        FROM franchise_health_snapshots
        GROUP BY year, month
        ORDER BY year DESC, month DESC
        LIMIT 12
    """)

    return render_template(
        "admin/business_intelligence.html",
        selected_year=selected_year,
        selected_month=selected_month,
        latest_period=latest,
        summary=summary,
        health_rows=health_rows,
        top_growth=top_growth,
        biggest_decline=biggest_decline,
        target_misses=target_misses,
        province_rows=province_rows,
        periods=periods,
    )


@admin_bp.route("/business-intelligence/rebuild", methods=["POST"])
@login_required
def business_intelligence_rebuild():
    if not can_view_operations_centre():
        abort(403)
    from app.business_intelligence import rebuild_business_intelligence, latest_period
    latest = latest_period()
    selected_year = int(request.form.get("year") or latest.get("year") or datetime.utcnow().year)
    selected_month = int(request.form.get("month") or latest.get("month") or datetime.utcnow().month)
    result = rebuild_business_intelligence(selected_year, selected_month, commit=True)
    try:
        from app.events import emit_event
        emit_event(
            "business_intelligence.rebuilt",
            source="business_intelligence",
            title="Business intelligence rebuilt",
            message=f"Business intelligence rebuilt for {selected_year}-{selected_month:02d}.",
            year=selected_year,
            month=selected_month,
            payload=result,
        )
    except Exception:
        db.session.rollback()
    log_action("Business Intelligence", "Rebuilt BI", f"Period {selected_year}-{selected_month:02d}; snapshots {result.get('snapshots', 0)}")
    flash(f"Business Intelligence rebuilt for {selected_year}-{selected_month:02d}.", "success")
    return redirect(url_for("admin.business_intelligence", year=selected_year, month=selected_month))


@admin_bp.route("/insights")
@login_required
def insights_dashboard():
    """Phase 12 Enterprise Insights and Explanation Engine dashboard."""
    if not can_view_operations_centre():
        abort(403)
    from app.insights_engine import latest_period, get_insight_summary
    latest = latest_period()
    selected_year = int(request.args.get("year") or latest.get("year") or datetime.utcnow().year)
    selected_month = int(request.args.get("month") or latest.get("month") or datetime.utcnow().month)
    active_tab = (request.args.get("tab") or "executive").strip().lower()
    allowed_tabs = {"executive", "franchise", "royalty", "province"}
    if active_tab not in allowed_tabs:
        active_tab = "executive"

    summary = get_insight_summary(selected_year, selected_month)

    # Older/stale narrative sets may contain only the executive record.  Rebuild
    # once when source data exists but one or more sub-tab datasets are absent.
    # This keeps all four tabs populated without rebuilding on every page view.
    narrative_counts = dict(
        db.session.query(InsightNarrative.narrative_type, db.func.count(InsightNarrative.id))
        .filter_by(year=selected_year, month=selected_month)
        .group_by(InsightNarrative.narrative_type)
        .all()
    )
    source_counts = {
        "franchise": FranchiseHealthSnapshot.query.filter_by(year=selected_year, month=selected_month).count(),
        "royalty": RoyaltyCalculationSnapshot.query.filter_by(year=selected_year, month=selected_month).count(),
    }
    expected_missing = (
        (source_counts["franchise"] > 0 and not (narrative_counts.get("franchise_performance") or narrative_counts.get("business_insight_explanation")))
        or (source_counts["royalty"] > 0 and not (narrative_counts.get("royalty_explanation") or narrative_counts.get("royalty_warning")))
        or (source_counts["franchise"] > 0 and not narrative_counts.get("province_summary"))
    )
    if expected_missing:
        from app.insights_engine import rebuild_insight_narratives
        rebuild_insight_narratives(selected_year, selected_month, commit=True)
        summary = get_insight_summary(selected_year, selected_month)

    narratives = (InsightNarrative.query.filter_by(year=selected_year, month=selected_month)
        .order_by(InsightNarrative.created_at.desc(), InsightNarrative.id.desc()).limit(240).all())

    # Keep each sub-tab tied to one narrative type.  Previously every record with
    # a franchise_id appeared under Franchise Explanations, which duplicated
    # royalty and BI information under different headings.
    executive_items = [n for n in narratives if n.narrative_type in ("executive_summary", "company_health", "monthly_summary")]
    franchise_items = [n for n in narratives if n.narrative_type in ("franchise_performance", "business_insight_explanation")][:40]
    province_items = [n for n in narratives if n.narrative_type == "province_summary"][:40]
    royalty_items = [n for n in narratives if n.narrative_type in ("royalty_explanation", "royalty_warning")][:40]
    return render_template(
        "admin/insights_dashboard.html",
        selected_year=selected_year,
        selected_month=selected_month,
        latest_period=latest,
        summary=summary,
        narratives=narratives,
        executive_items=executive_items,
        franchise_items=franchise_items,
        province_items=province_items,
        royalty_items=royalty_items,
        active_tab=active_tab,
    )


@admin_bp.route("/insights/rebuild", methods=["POST"])
@login_required
def insights_rebuild():
    if not can_view_operations_centre():
        abort(403)
    from app.insights_engine import rebuild_insight_narratives, latest_period
    latest = latest_period()
    selected_year = int(request.form.get("year") or latest.get("year") or datetime.utcnow().year)
    selected_month = int(request.form.get("month") or latest.get("month") or datetime.utcnow().month)
    active_tab = (request.form.get("tab") or "executive").strip().lower()
    if active_tab not in {"executive", "franchise", "royalty", "province"}:
        active_tab = "executive"
    result = rebuild_insight_narratives(selected_year, selected_month, commit=True)
    try:
        from app.events import emit_event
        emit_event(
            "insights.rebuilt",
            source="insights_engine",
            title="Insight narratives rebuilt",
            message=f"Insight narratives rebuilt for {selected_year}-{selected_month:02d}.",
            year=selected_year,
            month=selected_month,
            payload=result,
        )
    except Exception:
        db.session.rollback()
    log_action("Insights", "Rebuilt insight narratives", f"Period {selected_year}-{selected_month:02d}; narratives {result.get('narratives', 0)}")
    flash(f"Insight narratives rebuilt for {selected_year}-{selected_month:02d}.", "success")
    return redirect(url_for("admin.insights_dashboard", year=selected_year, month=selected_month, tab=active_tab))


@admin_bp.route("/operations")
@login_required
def operations_centre():
    """Enterprise Operations Centre for live system monitoring and recovery."""
    if not can_view_operations_centre():
        abort(403)

    import_status_rows = db.session.execute(text("""
        SELECT status, COUNT(*) AS count
        FROM import_jobs
        GROUP BY status
        ORDER BY status
    """)).mappings().all()
    import_status = {row["status"] or "unknown": row["count"] for row in import_status_rows}
    needs_review_imports = ImportJob.query.filter(ImportJob.status.in_(["needs_review", "failed", "warning"])).order_by(ImportJob.started_at.desc()).limit(12).all()
    running_imports = ImportJob.query.filter(ImportJob.status.in_(["queued", "running", "processing", "validating", "publishing"])).order_by(ImportJob.started_at.desc()).limit(12).all()
    stuck_jobs = ImportJob.query.filter(
        ImportJob.status.in_(["running", "processing", "validating", "publishing"]),
        ImportJob.heartbeat_at.isnot(None),
        ImportJob.heartbeat_at < (datetime.utcnow() - timedelta(minutes=15)),
    ).order_by(ImportJob.heartbeat_at.asc()).limit(12).all()
    latest_imports = ImportJob.query.order_by(ImportJob.started_at.desc()).limit(12).all()
    latest_job_logs = ImportJobLog.query.order_by(ImportJobLog.created_at.desc()).limit(20).all()
    worker_heartbeats = WorkerHeartbeat.query.order_by(WorkerHeartbeat.heartbeat_at.desc()).limit(12).all()
    online_workers = [worker for worker in worker_heartbeats if worker.is_online and worker.status != "stopped"]

    latest_period = _latest_import_period()
    monthly_period_rows = db.session.execute(text("""
        SELECT year, month, COUNT(*) AS rows, COALESCE(SUM(gross_turnover), 0) AS gross_turnover,
               COALESCE(SUM(royalty_amount), 0) AS royalty_amount, COALESCE(SUM(payover), 0) AS payover,
               MAX(updated_at) AS last_updated
        FROM monthly_figures
        GROUP BY year, month
        ORDER BY year DESC, month DESC
        LIMIT 6
    """)).mappings().all()

    diagnostics = {
        "missing_scales": db.session.execute(text("""
            SELECT COUNT(*) FROM franchises f
            LEFT JOIN royalty_scales rs ON rs.franchise_id = f.id
            GROUP BY f.id
            HAVING COUNT(rs.id) = 0
        """)).fetchall(),
        "missing_agreements": db.session.execute(text("""
            SELECT COUNT(*) FROM franchises
            WHERE agreement_start_date IS NULL OR agreement_end_date IS NULL
        """)).scalar() or 0,
        "orphan_monthly_figures": db.session.execute(text("""
            SELECT COUNT(*) FROM monthly_figures mf
            LEFT JOIN franchises f ON f.id = mf.franchise_id
            WHERE f.id IS NULL
        """)).scalar() or 0,
        "zero_royalty_warnings": db.session.execute(text("""
            SELECT COUNT(*) FROM monthly_figures
            WHERE COALESCE(gross_turnover, 0) > 0 AND COALESCE(royalty_amount, 0) = 0
        """)).scalar() or 0,
    }
    diagnostics["missing_scales"] = len(diagnostics["missing_scales"])

    cache = cache_stats()
    latest_cache_rows = PerformancePageCache.query.order_by(PerformancePageCache.built_at.desc()).limit(8).all()
    latest_events = LiveEvent.query.order_by(LiveEvent.created_at.desc()).limit(12).all()
    latest_notifications = LiveNotification.query.order_by(LiveNotification.created_at.desc()).limit(12).all()
    latest_audit_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(12).all()

    try:
        from app.events import event_stats, ensure_default_subscriptions
        ensure_default_subscriptions(commit=False)
        event_bus_stats = event_stats()
    except Exception:
        event_bus_stats = {}
    latest_system_events = SystemEvent.query.order_by(SystemEvent.created_at.desc()).limit(15).all()
    failed_system_events = SystemEvent.query.filter(SystemEvent.status.in_(["failed", "needs_review"])).order_by(SystemEvent.created_at.desc()).limit(10).all()
    event_subscriptions = EventSubscription.query.order_by(EventSubscription.event_type, EventSubscription.name).limit(20).all()
    latest_event_logs = EventProcessingLog.query.order_by(EventProcessingLog.created_at.desc()).limit(15).all()

    unread_notifications = LiveNotification.query.filter(LiveNotification.read_at.is_(None)).count()
    admin_cards = [
        {"label": "Latest period", "value": f"{latest_period['year']}-{int(latest_period['month']):02d}", "tone": "ok" if latest_period.get("rows") else "warning"},
        {"label": "Import jobs", "value": _ops_count("import_jobs"), "tone": "ok"},
        {"label": "Needs review", "value": sum(import_status.get(key, 0) for key in ("needs_review", "failed", "warning")), "tone": "danger" if sum(import_status.get(key, 0) for key in ("needs_review", "failed", "warning")) else "ok"},
        {"label": "Running imports", "value": len(running_imports), "tone": "warning" if running_imports else "ok"},
        {"label": "Stale jobs", "value": len(stuck_jobs), "tone": "danger" if stuck_jobs else "ok"},
        {"label": "Online workers", "value": len(online_workers), "tone": "ok" if online_workers else "warning"},
        {"label": "Valid cache rows", "value": cache.get("valid", 0), "tone": "ok" if cache.get("valid", 0) else "warning"},
        {"label": "Unread notifications", "value": unread_notifications, "tone": "warning" if unread_notifications else "ok"},
        {"label": "Pending events", "value": event_bus_stats.get("pending", 0), "tone": "warning" if event_bus_stats.get("pending", 0) else "ok"},
        {"label": "Failed events", "value": event_bus_stats.get("failed", 0), "tone": "danger" if event_bus_stats.get("failed", 0) else "ok"},
    ]

    health_checks = [
        {"label": "Database connection", "status": "OK", "tone": "ok", "detail": "The application can query PostgreSQL."},
        {"label": "Import pipeline", "status": "Review" if admin_cards[2]["value"] else "OK", "tone": "danger" if admin_cards[2]["value"] else "ok", "detail": f"{admin_cards[2]['value']} jobs require attention."},
        {"label": "Royalty calculation", "status": "Review" if diagnostics["zero_royalty_warnings"] else "OK", "tone": "warning" if diagnostics["zero_royalty_warnings"] else "ok", "detail": f"{diagnostics['zero_royalty_warnings']} gross-turnover rows have zero royalty."},
        {"label": "Franchise agreements", "status": "Review" if diagnostics["missing_agreements"] else "OK", "tone": "warning" if diagnostics["missing_agreements"] else "ok", "detail": f"{diagnostics['missing_agreements']} franchises have missing agreement dates."},
        {"label": "Performance cache", "status": "OK" if cache.get("valid", 0) else "Needs warmup", "tone": "ok" if cache.get("valid", 0) else "warning", "detail": f"{cache.get('valid', 0)} valid / {cache.get('invalidated', 0)} invalidated cache rows."},
        {"label": "Background worker", "status": "Online" if online_workers else "Not detected", "tone": "ok" if online_workers else "warning", "detail": f"{len(online_workers)} active worker heartbeat row(s)."},
        {"label": "Event bus", "status": "Review" if event_bus_stats.get("failed", 0) else "OK", "tone": "danger" if event_bus_stats.get("failed", 0) else "ok", "detail": f"{event_bus_stats.get('pending', 0)} pending / {event_bus_stats.get('failed', 0)} failed / {event_bus_stats.get('processed', 0)} processed events."},
    ]

    table_counts = [
        {"table": "users", "count": _ops_count("users")},
        {"table": "franchises", "count": _ops_count("franchises")},
        {"table": "monthly_figures", "count": _ops_count("monthly_figures")},
        {"table": "royalty_scales", "count": _ops_count("royalty_scales")},
        {"table": "import_jobs", "count": _ops_count("import_jobs")},
        {"table": "worker_heartbeats", "count": _ops_count("worker_heartbeats")},
        {"table": "live_events", "count": _ops_count("live_events")},
        {"table": "system_events", "count": _ops_count("system_events")},
        {"table": "event_processing_logs", "count": _ops_count("event_processing_logs")},
        {"table": "performance_page_cache", "count": _ops_count("performance_page_cache")},
        {"table": "audit_logs", "count": _ops_count("audit_logs")},
    ]

    return render_template(
        "admin/operations_centre.html",
        admin_cards=admin_cards,
        health_checks=health_checks,
        diagnostics=diagnostics,
        import_status=import_status,
        running_imports=running_imports,
        needs_review_imports=needs_review_imports,
        latest_imports=latest_imports,
        stuck_jobs=stuck_jobs,
        latest_job_logs=latest_job_logs,
        worker_heartbeats=worker_heartbeats,
        online_workers=online_workers,
        monthly_period_rows=monthly_period_rows,
        cache=cache,
        latest_cache_rows=latest_cache_rows,
        latest_events=latest_events,
        latest_notifications=latest_notifications,
        latest_audit_logs=latest_audit_logs,
        event_bus_stats=event_bus_stats,
        latest_system_events=latest_system_events,
        failed_system_events=failed_system_events,
        event_subscriptions=event_subscriptions,
        latest_event_logs=latest_event_logs,
        table_counts=table_counts,
        latest_period=latest_period,
    )


@admin_bp.route("/operations/cache/invalidate", methods=["POST"])
@login_required
def operations_invalidate_cache():
    if not can_view_operations_centre():
        abort(403)
    count = invalidate_performance_cache(commit=True)
    log_action("Operations", "Invalidated performance cache", f"Rows invalidated: {count}")
    flash(f"Performance cache invalidated ({count} rows). It will rebuild after the next import or graph request.", "success")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/cache/rebuild", methods=["POST"])
@login_required
def operations_rebuild_cache():
    if not can_view_operations_centre():
        abort(403)
    month = request.form.get("month", type=int)
    year = request.form.get("year", type=int)
    if not month or not year:
        latest = _latest_import_period()
        month = int(latest.get("month") or datetime.utcnow().month)
        year = int(latest.get("year") or datetime.utcnow().year)
    franchise_ids = [fid for (fid,) in db.session.query(Franchise.id).filter(Franchise.is_performance_active == True).all()]
    result = warm_performance_cache_for_period(month, year, franchise_ids=franchise_ids)
    try:
        from app.events import emit_event
        emit_event(
            "cache.rebuilt",
            source="operations.cache.rebuild",
            title=f"Performance cache rebuilt for {year}-{month:02d}",
            message=f"{result.get('cache_rows', 0)} cache rows rebuilt.",
            payload=result,
            year=year,
            month=month,
            aggregate_type="performance_cache",
        )
    except Exception:
        pass
    log_action("Operations", "Rebuilt performance cache", f"Period: {year}-{month:02d}; Result: {result}")
    flash(f"Performance cache rebuilt for {year}-{month:02d}: {result.get('cache_rows', 0)} cache rows and {result.get('performance_rows', 0)} performance rows.", "success")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/jobs/run-next", methods=["POST"])
@login_required
def operations_run_next_job():
    if not can_view_operations_centre():
        abort(403)
    from app.jobs import run_next_job
    job = run_next_job(worker_id=f"web:{current_user.id}")
    if job:
        log_action("Operations", "Processed queued job", f"Job {job.id}: {job.status} - {job.message}")
        flash(f"Processed job {job.id}: {job.status} - {job.message}", "success" if job.status == "completed" else "warning")
    else:
        flash("No queued jobs found.", "info")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/jobs/<int:job_id>/retry", methods=["POST"])
@login_required
def operations_retry_job(job_id):
    if not can_view_operations_centre():
        abort(403)
    from app.jobs import retry_job
    job = ImportJob.query.get_or_404(job_id)
    retry_job(job)
    log_action("Operations", "Retried job", f"Job {job.id}: {job.kind}")
    flash(f"Job {job.id} queued for retry.", "success")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/jobs/<int:job_id>/cancel", methods=["POST"])
@login_required
def operations_cancel_job(job_id):
    if not can_view_operations_centre():
        abort(403)
    from app.jobs import cancel_job
    job = ImportJob.query.get_or_404(job_id)
    cancel_job(job, reason=f"Cancelled by {current_user.email}")
    log_action("Operations", "Cancelled job", f"Job {job.id}: {job.kind}")
    flash(f"Job {job.id} cancelled.", "warning")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/jobs/release-stale", methods=["POST"])
@login_required
def operations_release_stale_jobs():
    if not can_view_operations_centre():
        abort(403)
    from app.jobs import release_stale_jobs
    count = release_stale_jobs(stale_after_minutes=15, worker_id=f"web:{current_user.id}")
    log_action("Operations", "Released stale jobs", f"Released stale jobs: {count}")
    flash(f"Released stale jobs: {count}", "success" if count else "info")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/jobs/<int:job_id>/release", methods=["POST"])
@login_required
def operations_release_job(job_id):
    if not can_view_operations_centre():
        abort(403)
    job = ImportJob.query.get_or_404(job_id)
    job.status = "queued"
    job.locked_at = None
    job.locked_by = None
    job.message = "Released back to queue by Admin."
    db.session.commit()
    log_action("Operations", "Released stale job", f"Job {job.id}: {job.kind}")
    flash(f"Job {job.id} released back to the queue.", "success")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/events/process", methods=["POST"])
@login_required
def operations_process_events():
    if not can_view_operations_centre():
        abort(403)
    from app.events import process_pending_events
    count = process_pending_events(limit=50, worker_id=f"web:{current_user.id}")
    log_action("Operations", "Processed event bus", f"Events processed: {count}")
    flash(f"Processed event bus events: {count}", "success" if count else "info")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/events/<int:event_id>/retry", methods=["POST"])
@login_required
def operations_retry_event(event_id):
    if not can_view_operations_centre():
        abort(403)
    from app.events import retry_event
    event = SystemEvent.query.get_or_404(event_id)
    retry_event(event)
    log_action("Operations", "Retried event", f"Event {event.id}: {event.event_type}")
    flash(f"Event {event.id} queued for retry.", "success")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/operations/events/release-stale", methods=["POST"])
@login_required
def operations_release_stale_events():
    if not can_view_operations_centre():
        abort(403)
    from app.events import release_stale_events
    count = release_stale_events(stale_after_minutes=15, worker_id=f"web:{current_user.id}")
    log_action("Operations", "Released stale events", f"Events released: {count}")
    flash(f"Released stale events: {count}", "success" if count else "info")
    return redirect(url_for("admin.operations_centre"))


@admin_bp.route("/roles/new", methods=["GET", "POST"])
@login_required
@permission_required("user_roles:add")
def new_role():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        if not name:
            flash("Role name is required.", "danger")
            return render_template("admin/new_role.html")
        if Role.query.filter_by(name=name).first():
            flash("A role with this name already exists.", "warning")
            return render_template("admin/new_role.html")
        role = Role(name=name, description=description)
        db.session.add(role)
        log_action("User Roles", "Created role", f"Role: {role.name}")
        db.session.commit()
        flash("Role created. You can now tick its permissions.", "success")
        return redirect(url_for("admin.edit_role", role_id=role.id))
    return render_template("admin/new_role.html")






@admin_bp.route("/royalty-management")
@login_required
def royalty_management():
    """Admin-only Enterprise Royalty Management dashboard."""
    if not can_view_operations_centre():
        abort(403)
    from app.royalty_management import royalty_management_summary
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    if not month or not year:
        latest = _latest_import_period()
        month = int(latest.get("month") or datetime.utcnow().month)
        year = int(latest.get("year") or datetime.utcnow().year)
    summary = royalty_management_summary()
    period_snapshots = (RoyaltyCalculationSnapshot.query
        .filter_by(month=month, year=year)
        .order_by(RoyaltyCalculationSnapshot.status.desc(), RoyaltyCalculationSnapshot.royalty_amount.desc())
        .limit(200).all())
    needs_review = [item for item in period_snapshots if item.status == "needs_review"]
    period_rows = db.session.execute(text("""
        SELECT year, month, COUNT(*) AS rows, COALESCE(SUM(royalty_amount), 0) AS royalty_amount,
               COALESCE(SUM(gross_turnover), 0) AS gross_turnover
        FROM monthly_figures
        GROUP BY year, month
        ORDER BY year DESC, month DESC
        LIMIT 18
    """)).mappings().all()
    return render_template(
        "admin/royalty_management.html",
        selected_month=month,
        selected_year=year,
        summary=summary,
        period_snapshots=period_snapshots,
        needs_review=needs_review,
        period_rows=period_rows,
        growth_profiles=summary.get("growth_profiles", []),
    )


@admin_bp.route("/royalty-management/recalculate", methods=["POST"])
@login_required
def royalty_management_recalculate():
    """Rebuild royalty snapshots and existing royalty amounts for one period."""
    if not can_view_operations_centre():
        abort(403)
    from app.royalty_management import recalculate_royalties_for_period
    month = request.form.get("month", type=int)
    year = request.form.get("year", type=int)
    if not month or not year:
        latest = _latest_import_period()
        month = int(latest.get("month") or datetime.utcnow().month)
        year = int(latest.get("year") or datetime.utcnow().year)
    result = recalculate_royalties_for_period(month, year, commit=True)
    log_action("Royalty Management", "Recalculated royalties", f"Period {year}-{month:02d}: {result}")
    if result.get("needs_review"):
        flash(f"Royalties recalculated for {year}-{month:02d}. {result['calculated']} calculated, {result['needs_review']} need review.", "warning")
    else:
        flash(f"Royalties recalculated for {year}-{month:02d}. {result['calculated']} rows calculated successfully.", "success")
    return redirect(url_for("admin.royalty_management", month=month, year=year))


@admin_bp.route("/royalty-management/growth-profile/<int:profile_id>", methods=["POST"])
@login_required
def royalty_management_update_growth_profile(profile_id):
    """Admin-only GDP/growth standard update."""
    if not is_current_user_admin():
        abort(403)
    profile = RoyaltyGrowthProfile.query.get_or_404(profile_id)
    old = str(profile.default_growth_percent)
    raw = (request.form.get("default_growth_percent") or "").replace(",", ".").strip()
    try:
        value = float(raw)
    except ValueError:
        flash("Invalid growth percentage.", "danger")
        return redirect(url_for("admin.royalty_management"))
    profile.default_growth_percent = value
    profile.notes = request.form.get("notes", profile.notes) or profile.notes
    db.session.add(RoyaltyOverride(
        franchise_id=None,
        override_type="growth_profile",
        field_name="default_growth_percent",
        old_value=old,
        new_value=str(value),
        reason=request.form.get("reason", "Admin updated global growth profile") or "Admin updated global growth profile",
        created_by_id=current_user.id,
    ))
    db.session.commit()
    log_action("Royalty Management", "Updated growth profile", f"{profile.name}: {old} -> {value}")
    flash("Growth profile updated. Recalculate royalties for the affected periods when you are ready.", "success")
    return redirect(url_for("admin.royalty_management"))


@admin_bp.route("/database-diagnostics")
@login_required
def database_diagnostics():
    """Admin diagnostics page for import, franchise, royalty and user-link health checks."""
    if not (is_current_user_admin() or current_user.has_role("Finance Manager") or current_user.has_permission("system_administration:view")):
        abort(403)

    # Franchises that cannot calculate royalties because no scale rows exist.
    missing_scales = db.session.execute(text("""
        SELECT f.id, f.business_name, f.franchise_code, f.agreement_start_date, f.royalty_gross_method, COUNT(rs.id) AS scale_count
        FROM franchises f
        LEFT JOIN royalty_scales rs ON rs.franchise_id = f.id
        GROUP BY f.id, f.business_name, f.franchise_code, f.agreement_start_date, f.royalty_gross_method
        HAVING COUNT(rs.id) = 0
        ORDER BY f.business_name
        LIMIT 250
    """)).mappings().all()

    missing_agreements = db.session.execute(text("""
        SELECT id, business_name, franchise_code, agreement_start_date, agreement_end_date
        FROM franchises
        WHERE agreement_start_date IS NULL
           OR agreement_end_date IS NULL
        ORDER BY business_name
        LIMIT 250
    """)).mappings().all()

    franchise_users_without_links = db.session.execute(text("""
        SELECT u.id, u.name, u.surname, u.email, string_agg(r.name, ', ' ORDER BY r.name) AS roles
        FROM users u
        JOIN user_roles ur ON ur.user_id = u.id
        JOIN roles r ON r.id = ur.role_id
        LEFT JOIN user_franchises uf ON uf.user_id = u.id
        WHERE r.name = 'Franchise User'
        GROUP BY u.id, u.name, u.surname, u.email
        HAVING COUNT(uf.franchise_id) = 0
        ORDER BY u.email
        LIMIT 250
    """)).mappings().all()

    franchise_employees_without_parent = db.session.execute(text("""
        SELECT u.id, u.name, u.surname, u.email, string_agg(r.name, ', ' ORDER BY r.name) AS roles
        FROM users u
        JOIN user_roles ur ON ur.user_id = u.id
        JOIN roles r ON r.id = ur.role_id
        WHERE r.name IN ('Franchise Manager', 'Franchise Employee', 'Franchise Agent')
          AND u.parent_franchise_user_id IS NULL
        GROUP BY u.id, u.name, u.surname, u.email
        ORDER BY u.email
        LIMIT 250
    """)).mappings().all()

    duplicate_franchises = db.session.execute(text("""
        SELECT lower(trim(business_name)) AS normalized_name, COUNT(*) AS count, string_agg(id::text, ', ' ORDER BY id) AS franchise_ids
        FROM franchises
        WHERE business_name IS NOT NULL AND trim(business_name) <> ''
        GROUP BY lower(trim(business_name))
        HAVING COUNT(*) > 1
        ORDER BY count DESC, normalized_name
        LIMIT 250
    """)).mappings().all()

    latest_imports = ImportJob.query.order_by(ImportJob.started_at.desc()).limit(20).all()

    orphan_monthly_figures = db.session.execute(text("""
        SELECT mf.id, mf.year, mf.month, mf.franchise_id, mf.gross_turnover, mf.royalty_amount, mf.payover
        FROM monthly_figures mf
        LEFT JOIN franchises f ON f.id = mf.franchise_id
        WHERE f.id IS NULL
        ORDER BY mf.year DESC, mf.month DESC, mf.id DESC
        LIMIT 250
    """)).mappings().all()

    latest_month = db.session.execute(text("""
        SELECT year, month, COUNT(*) AS rows, COALESCE(SUM(gross_turnover), 0) AS gross_turnover,
               COALESCE(SUM(royalty_amount), 0) AS royalty_amount, COALESCE(SUM(payover), 0) AS payover
        FROM monthly_figures
        GROUP BY year, month
        ORDER BY year DESC, month DESC
        LIMIT 12
    """)).mappings().all()

    royalty_zero_checks = db.session.execute(text("""
        SELECT mf.year, mf.month, f.business_name, mf.franchise_id, mf.gross_turnover, mf.royalty_percentage, mf.royalty_amount, mf.payover
        FROM monthly_figures mf
        JOIN franchises f ON f.id = mf.franchise_id
        WHERE COALESCE(mf.gross_turnover, 0) > 0
          AND COALESCE(mf.royalty_amount, 0) = 0
        ORDER BY mf.year DESC, mf.month DESC, f.business_name
        LIMIT 250
    """)).mappings().all()

    performance_cache_stats = cache_stats()
    latest_cache_rows = PerformancePageCache.query.order_by(PerformancePageCache.built_at.desc()).limit(20).all()

    summary_cards = [
        {"label": "Missing royalty scales", "value": len(missing_scales), "tone": "danger" if missing_scales else "ok"},
        {"label": "Missing agreement dates", "value": len(missing_agreements), "tone": "warning" if missing_agreements else "ok"},
        {"label": "Franchise users not linked", "value": len(franchise_users_without_links), "tone": "danger" if franchise_users_without_links else "ok"},
        {"label": "Employees not linked", "value": len(franchise_employees_without_parent), "tone": "warning" if franchise_employees_without_parent else "ok"},
        {"label": "Orphan monthly figures", "value": len(orphan_monthly_figures), "tone": "danger" if orphan_monthly_figures else "ok"},
        {"label": "Zero royalty warnings", "value": len(royalty_zero_checks), "tone": "warning" if royalty_zero_checks else "ok"},
        {"label": "Valid performance cache", "value": performance_cache_stats.get("valid", 0), "tone": "ok" if performance_cache_stats.get("valid", 0) else "warning"},
    ]

    return render_template(
        "admin/database_diagnostics.html",
        summary_cards=summary_cards,
        missing_scales=missing_scales,
        missing_agreements=missing_agreements,
        franchise_users_without_links=franchise_users_without_links,
        franchise_employees_without_parent=franchise_employees_without_parent,
        duplicate_franchises=duplicate_franchises,
        latest_imports=latest_imports,
        orphan_monthly_figures=orphan_monthly_figures,
        latest_month=latest_month,
        royalty_zero_checks=royalty_zero_checks,
        performance_cache_stats=performance_cache_stats,
        latest_cache_rows=latest_cache_rows,
    )


@admin_bp.route("/audit-logs")
@login_required
@permission_required("audit_logs:view")
def audit_logs():
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(300).all()
    return render_template("admin/audit_logs.html", logs=logs)

@admin_bp.route("/users/<int:user_id>/activate", methods=["POST"])
@login_required
@permission_required("users:edit")
def activate_user(user_id):
    user = User.query.get_or_404(user_id)

    user.is_active = True

    db.session.commit()

    flash(f"{user.name} activated.", "success")

    return redirect(url_for("admin.users"))

@admin_bp.route("/users/<int:user_id>/franchises", methods=["POST"])
@login_required
@permission_required("users:edit")
def assign_user_franchises(user_id):
    if not can_assign_franchise_links():
        flash("Your role does not have permission to link franchise users to franchises.", "danger")
        return redirect(url_for("admin.users"))

    user = User.query.get_or_404(user_id)
    if is_admin_side_user(user):
        user.assigned_franchises = []
        db.session.commit()
        flash("Martins Funerals South Africa/admin-side users are not linked to franchises here.", "warning")
        return redirect(url_for("admin.users"))

    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids")]
    user.assigned_franchises = Franchise.query.filter(Franchise.id.in_(franchise_ids)).all() if franchise_ids else []
    log_action("Users", "Updated user franchise access", f"User: {user.full_name}")
    db.session.commit()
    flash(f"Franchise access updated for {user.full_name}.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/clean-finance-admin-users", methods=["POST"])
@login_required
@permission_required("users:edit")
def clean_finance_admin_users():
    changed = tidy_finance_admin_users()
    log_action("Users", "Cleaned finance/admin-side users", f"Changes: {changed}")
    db.session.commit()
    flash("Renette, Lowhaan and Deon were cleaned as Martins Funerals South Africa/admin-side users with no franchise links.", "success")
    return redirect(url_for("admin.users"))

@admin_bp.route("/users/clear-franchise-user-links", methods=["POST"])
@login_required
@permission_required("users:edit")
def clear_franchise_user_links():
    if not can_assign_franchise_links():
        flash("Your role does not have permission to clear franchise user links.", "danger")
        return redirect(url_for("admin.users"))

    target_role_names = {"Franchise User", "Franchise Manager", "Regional Manager"}
    protected_role_names = {"Admin", "Finance Manager", "Finance Assistant"}
    cleared = 0

    users = User.query.all()
    for user in users:
        role_names = {role.name for role in user.roles}
        if role_names & target_role_names and not (role_names & protected_role_names):
            if user.assigned_franchises:
                user.assigned_franchises = []
                cleared += 1

    log_action("Users", "Cleared franchise user links", f"Users cleared: {cleared}")
    db.session.commit()
    flash(f"Cleared linked franchises from {cleared} franchise user(s). Finance and admin users were not changed.", "success")
    return redirect(url_for("admin.users"))

# ---------------------------------------------------------------------------
# Admin oversight for employee users created by franchise users
# ---------------------------------------------------------------------------

def is_franchise_employee_user(user):
    """Return True for any employee account that belongs under a franchise user.

    Older records were not always saved with parent_franchise_user_id, so Admin
    must also treat the franchise-side employee roles as employee accounts. This
    keeps Manager, Employee and Agent accounts visible/editable in Admin >
    Employees.
    """
    franchise_owner_ids = {owner.id for owner in all_franchise_owner_users()}
    return (
        user.has_role("Franchise Manager")
        or user.has_role("Franchise Employee")
        or user.has_role("Franchise Agent")
        or bool(getattr(user, "parent_franchise_user_id", None))
        or bool(getattr(user, "created_by_user_id", None) in franchise_owner_ids)
    )


@admin_bp.route("/franchise-employees")
@login_required
@permission_required("users:view")
def franchise_employees():
    repair_existing_user_visibility()
    db.session.commit()
    employees = [user for user in User.query.order_by(User.name, User.surname).all() if is_franchise_employee_user(user)]
    owner_ids = [employee.parent_franchise_user_id for employee in employees if employee.parent_franchise_user_id]
    owners = {user.id: user for user in User.query.filter(User.id.in_(owner_ids)).all()} if owner_ids else {}
    now = datetime.utcnow()
    franchise_owners = all_franchise_owner_users()
    franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
    employee_roles = Role.query.filter(Role.name.in_(["Franchise Manager", "Franchise Employee", "Franchise Agent"])).order_by(Role.name).all()
    return render_template(
        "admin/franchise_employees.html",
        employees=employees,
        owners=owners,
        franchise_owners=franchise_owners,
        franchises=franchises,
        employee_roles=employee_roles,
    )


@admin_bp.route("/franchise-employees/create", methods=["POST"])
@login_required
@permission_required("users:add")
def create_franchise_employee_admin():
    name = request.form.get("name", "").strip()
    surname = request.form.get("surname", "").strip()
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "").strip()
    role_id = request.form.get("role_id", type=int)
    franchise_id = request.form.get("franchise_id", type=int)
    owner_id = request.form.get("parent_franchise_user_id", type=int)

    if not name or not surname or not email or not password or not role_id or not franchise_id:
        flash("Name, surname, email, password, role and franchise are required.", "danger")
        return redirect(url_for("admin.franchise_employees"))
    existing = User.query.filter(db.func.lower(User.email) == email).first()
    if existing:
        existing_roles = ", ".join(role.name for role in existing.roles) or "No role"
        flash(f"Email {email} already exists as {existing.full_name} ({existing_roles}). Use Edit to reset the password or activate the existing account.", "danger")
        return redirect(url_for("admin.franchise_employees"))

    selected_role = Role.query.get(role_id)
    if not selected_role or selected_role.name not in {"Franchise Manager", "Franchise Employee", "Franchise Agent"}:
        flash("Please select Manager, Employee or Agent.", "danger")
        return redirect(url_for("admin.franchise_employees"))

    franchise = Franchise.query.get_or_404(franchise_id)
    owner = User.query.get(owner_id) if owner_id else None
    if owner and not owner.has_role("Franchise User"):
        owner = None
    if not owner:
        owner = next((candidate for candidate in all_franchise_owner_users() if franchise in candidate.assigned_franchises), None)

    user = User(
        name=name,
        surname=surname,
        email=email,
        is_active=True,
        is_active_account=True,
        parent_franchise_user_id=owner.id if owner else None,
        created_by_user_id=current_user.id,
    )
    user.set_password(password)
    user.roles.append(selected_role)
    user.assigned_franchises.append(franchise)
    db.session.add(user)
    log_action("Franchise Employees", "Admin created franchise employee user", f"Employee: {email}; Franchise: {franchise.business_name}")
    db.session.commit()
    flash(f"Employee user {user.full_name} was created.", "success")
    return redirect(url_for("admin.franchise_employees"))


@admin_bp.route("/franchise-employees/<int:user_id>/update", methods=["POST"])
@login_required
@permission_required("users:edit")
def update_franchise_employee(user_id):
    user = User.query.get_or_404(user_id)
    if not is_franchise_employee_user(user):
        flash("This user is not a franchise employee user.", "danger")
        return redirect(url_for("admin.franchise_employees"))

    user.name = request.form.get("name", user.name).strip() or user.name
    user.surname = request.form.get("surname", user.surname).strip() or user.surname

    role_id = request.form.get("role_id", type=int)
    if role_id:
        selected_role = Role.query.get(role_id)
        if not selected_role or selected_role.name not in {"Franchise Manager", "Franchise Employee", "Franchise Agent"}:
            flash("Please select Manager, Employee or Agent for franchise employees.", "danger")
            return redirect(url_for("admin.franchise_employees"))
        user.roles = [selected_role]

    password = request.form.get("password", "").strip()
    if password:
        user.set_password(password)
    user.is_active = request.form.get("is_active") == "1"

    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids")]
    user.assigned_franchises = Franchise.query.filter(Franchise.id.in_(franchise_ids)).all() if franchise_ids else []

    owner_id = request.form.get("parent_franchise_user_id", type=int)
    if owner_id:
        owner = User.query.get(owner_id)
        if owner and owner != user:
            user.parent_franchise_user_id = owner.id

    log_action("Franchise Employees", "Admin updated franchise employee user", f"Employee: {user.email}")
    db.session.commit()
    flash(f"Employee user {user.full_name} was updated.", "success")
    return redirect(url_for("admin.franchise_employees"))


@admin_bp.route("/franchise-employees/<int:user_id>/delete", methods=["POST"])
@login_required
@permission_required("users:delete")
def delete_franchise_employee(user_id):
    user = User.query.get_or_404(user_id)
    if not is_franchise_employee_user(user):
        flash("This user is not a franchise employee user.", "danger")
        return redirect(url_for("admin.franchise_employees"))
    user.is_active = False
    user.is_active_account = False
    user.deactivated_at = datetime.utcnow()
    user.deactivation_reason = "Deactivated by Admin"
    log_action("Franchise Employees", "Admin deactivated franchise employee user", f"Employee: {user.email}")
    db.session.commit()
    flash(f"Employee user {user.full_name} was deactivated.", "success")
    return redirect(url_for("admin.franchise_employees"))


@admin_bp.route("/enterprise-workflows")
@login_required
def enterprise_workflows():
    """Phase 13 Enterprise Workflow & Automation Suite."""
    if not can_view_operations_centre():
        abort(403)
    from app.workflow_engine import ensure_phase13_defaults, workflow_summary
    ensure_phase13_defaults(commit=True)

    summary = workflow_summary()
    workflow_definitions = WorkflowDefinition.query.order_by(WorkflowDefinition.module, WorkflowDefinition.name).all()
    latest_workflows = WorkflowInstance.query.order_by(WorkflowInstance.created_at.desc()).limit(20).all()
    open_tasks = EnterpriseTask.query.filter_by(status="open").order_by(EnterpriseTask.priority.desc(), EnterpriseTask.created_at.desc()).limit(20).all()
    recent_notifications = EnterpriseNotification.query.order_by(EnterpriseNotification.created_at.desc()).limit(20).all()
    business_rules = BusinessRule.query.order_by(BusinessRule.module, BusinessRule.name).all()
    schedules = ScheduledJobDefinition.query.order_by(ScheduledJobDefinition.name).all()
    timeline = EnterpriseAuditTimeline.query.order_by(EnterpriseAuditTimeline.created_at.desc()).limit(30).all()

    workflow_status_rows = db.session.execute(text("""
        SELECT status, COUNT(*) AS count
        FROM workflow_instances
        GROUP BY status
        ORDER BY status
    """)).mappings().all()
    workflow_status = {row["status"]: row["count"] for row in workflow_status_rows}

    task_status_rows = db.session.execute(text("""
        SELECT status, COUNT(*) AS count
        FROM enterprise_tasks
        GROUP BY status
        ORDER BY status
    """)).mappings().all()
    task_status = {row["status"]: row["count"] for row in task_status_rows}

    return render_template(
        "admin/enterprise_workflows.html",
        summary=summary,
        workflow_definitions=workflow_definitions,
        latest_workflows=latest_workflows,
        open_tasks=open_tasks,
        recent_notifications=recent_notifications,
        business_rules=business_rules,
        schedules=schedules,
        timeline=timeline,
        workflow_status=workflow_status,
        task_status=task_status,
    )


@admin_bp.route("/enterprise-workflows/seed", methods=["POST"])
@login_required
def enterprise_workflows_seed():
    if not can_view_operations_centre():
        abort(403)
    from app.workflow_engine import ensure_phase13_defaults
    created = ensure_phase13_defaults(commit=True)
    log_action("Enterprise Workflows", "Seeded workflow defaults", str(created))
    flash(f"Workflow defaults ready: {created}", "success")
    return redirect(url_for("admin.enterprise_workflows"))


@admin_bp.route("/enterprise-workflows/run-diagnostics", methods=["POST"])
@login_required
def enterprise_workflows_run_diagnostics():
    if not can_view_operations_centre():
        abort(403)
    from app.workflow_engine import run_diagnostics_workflow
    instance = run_diagnostics_workflow(user_id=current_user.id, commit=True)
    log_action("Enterprise Workflows", "Ran diagnostics workflow", f"Workflow {instance.id}: {instance.status}")
    flash(f"Diagnostics workflow {instance.id} completed with status: {instance.status}", "success" if instance.status == "completed" else "warning")
    return redirect(url_for("admin.enterprise_workflows"))


@admin_bp.route("/enterprise-workflows/start/<workflow_key>", methods=["POST"])
@login_required
def enterprise_workflows_start(workflow_key):
    if not can_view_operations_centre():
        abort(403)
    from app.workflow_engine import start_workflow
    definition = WorkflowDefinition.query.filter_by(workflow_key=workflow_key).first_or_404()
    instance = start_workflow(
        workflow_key,
        title=f"Manual {definition.name}",
        module=definition.module,
        user_id=current_user.id,
        context={"source": "manual", "started_by": current_user.email},
        commit=True,
    )
    log_action("Enterprise Workflows", "Started workflow", f"Workflow {instance.id}: {workflow_key}")
    flash(f"Workflow started: {instance.title}", "success")
    return redirect(url_for("admin.enterprise_workflows"))


@admin_bp.route("/enterprise-workflows/tasks/<int:task_id>/complete", methods=["POST"])
@login_required
def enterprise_task_complete(task_id):
    if not can_view_operations_centre():
        abort(403)
    from app.workflow_engine import emit_timeline
    task = EnterpriseTask.query.get_or_404(task_id)
    task.status = "completed"
    task.completed_at = datetime.utcnow()
    emit_timeline(task.module, "task.completed", task.title, f"Completed by {current_user.email}", "info", user_id=current_user.id, workflow_instance_id=task.workflow_instance_id, franchise_id=task.franchise_id)
    db.session.commit()
    log_action("Enterprise Tasks", "Completed task", f"Task {task.id}: {task.title}")
    flash("Task completed.", "success")
    return redirect(url_for("admin.enterprise_workflows"))


@admin_bp.route("/enterprise-workflows/rules/<int:rule_id>/toggle", methods=["POST"])
@login_required
def enterprise_rule_toggle(rule_id):
    if not can_view_operations_centre():
        abort(403)
    rule = BusinessRule.query.get_or_404(rule_id)
    rule.is_active = not rule.is_active
    db.session.commit()
    log_action("Business Rules", "Toggled rule", f"{rule.rule_key}: {rule.is_active}")
    flash(f"Rule {'enabled' if rule.is_active else 'disabled'}: {rule.name}", "success")
    return redirect(url_for("admin.enterprise_workflows"))


@admin_bp.route("/enterprise-workflows/notifications/<int:notification_id>/read", methods=["POST"])
@login_required
def enterprise_notification_read(notification_id):
    if not can_view_operations_centre():
        abort(403)
    note = EnterpriseNotification.query.get_or_404(notification_id)
    note.is_read = True
    note.read_at = datetime.utcnow()
    db.session.commit()
    flash("Notification marked as read.", "success")
    return redirect(url_for("admin.enterprise_workflows"))


@admin_bp.route("/data-integrity")
@admin_bp.route("/franchise-master")
@login_required
def data_integrity():
    """Printable data integrity report for franchise master and royalty readiness.

    This page must never be blank. If a schema or data problem is found it renders
    a clear repair panel instead of failing with a server error.
    """
    from app.franchise_master_data import data_integrity_rows, get_latest_period, has_required_schema

    if not (current_user.has_role("Admin") or current_user.has_role("Finance Manager") or current_user.has_role("Finance Assistant")):
        abort(403)
    page_error = None
    schema_ok, missing_schema = has_required_schema()
    rows = []
    latest_year = latest_month = None
    if not schema_ok:
        page_error = "Missing database columns: " + ", ".join(missing_schema)
    else:
        try:
            rows = data_integrity_rows()
            latest_year, latest_month = get_latest_period()
        except Exception as exc:
            current_app.logger.exception("Data Integrity page failed")
            page_error = str(exc)
    if latest_year is None or latest_month is None:
        latest_year, latest_month = get_latest_period()
    summary = {
        "total": len(rows),
        "ready": sum(1 for r in rows if r["status"] == "Ready"),
        "needs_review": sum(1 for r in rows if r["status"] != "Ready"),
        "unassigned_regions": sum(1 for r in rows if not r.get("province") or r.get("province") == "Unassigned"),
        "missing_scales": sum(1 for r in rows if r.get("scale_count", 0) == 0),
        "missing_codes": sum(1 for r in rows if not r.get("franchise_code")),
    }
    return render_template(
        "admin/data_integrity.html",
        rows=rows,
        summary=summary,
        latest_year=latest_year,
        latest_month=latest_month,
        page_error=page_error,
        schema_ok=schema_ok,
    )




@admin_bp.route("/data-integrity/franchise/<int:franchise_id>/edit", methods=["GET", "POST"])
@admin_bp.route("/franchise-master/<int:franchise_id>/edit", methods=["GET", "POST"])
@login_required
def edit_franchise_master_record(franchise_id):
    """Edit one live franchise master record and its linked franchise user."""
    from decimal import Decimal, InvalidOperation
    if not (current_user.has_role("Admin") or current_user.has_role("Finance Manager") or current_user.has_role("Finance Assistant")):
        abort(403)
    franchise = Franchise.query.get_or_404(franchise_id)
    linked_user = (User.query.join(user_franchises, User.id == user_franchises.c.user_id)
                   .filter(user_franchises.c.franchise_id == franchise.id)
                   .order_by(user_franchises.c.is_primary.desc(), User.id.asc()).first())
    scales = RoyaltyScale.query.filter_by(franchise_id=franchise.id).order_by(RoyaltyScale.row_number.asc()).all()
    if request.method == "POST":
        def text_value(name): return (request.form.get(name) or "").strip()
        franchise.business_name = text_value("business_name")
        franchise.franchise_code = text_value("franchise_code").upper()
        franchise.master_import_id = text_value("master_import_id")
        franchise.standardized_town = text_value("standardized_town")
        franchise.province = text_value("province")
        franchise.province_code = text_value("province_code").upper()
        franchise.region = text_value("region")
        franchise.district = text_value("district")
        franchise.district_code = text_value("district_code").upper()
        franchise.municipality = text_value("municipality")
        franchise.municipality_code = text_value("municipality_code").upper()
        franchise.office_address = text_value("office_address")
        franchise.office_number = text_value("office_number")
        franchise.after_hours_number = text_value("after_hours_number")
        franchise.franchisee_name = text_value("franchisee_name")
        franchise.franchisee_surname = text_value("franchisee_surname")
        franchise.franchisee_cell = text_value("franchisee_cell")
        franchise.franchisee_email = text_value("franchisee_email").lower()
        franchise.public_email = text_value("public_email").lower()
        franchise.royalty_gross_method = text_value("royalty_gross_method") or "old"
        try: franchise.minimum_royalty_amount = Decimal(text_value("minimum_royalty_amount") or "0")
        except InvalidOperation:
            flash("Minimum royalty must be a valid number.", "danger")
            return render_template("admin/edit_franchise_master_record.html", franchise=franchise, linked_user=linked_user, scales=scales)
        for attr, field in (("agreement_start_date", "agreement_start_date"), ("agreement_end_date", "agreement_end_date")):
            raw = text_value(field)
            setattr(franchise, attr, datetime.strptime(raw, "%Y-%m-%d").date() if raw else None)
        login_email = text_value("login_email").lower()
        if login_email:
            duplicate = User.query.filter(db.func.lower(User.email) == login_email, User.id != (linked_user.id if linked_user else 0)).first()
            if duplicate:
                flash("That login email already belongs to another user.", "danger")
                return render_template("admin/edit_franchise_master_record.html", franchise=franchise, linked_user=linked_user, scales=scales)
            if linked_user:
                linked_user.email = login_email
                linked_user.name = text_value("login_name") or linked_user.name
                linked_user.surname = text_value("login_surname") or linked_user.surname
            else:
                role = Role.query.filter_by(name="Franchise User").first()
                if not role:
                    role = Role(name="Franchise User", description="Franchise owner/user", is_system_role=True)
                    db.session.add(role); db.session.flush()
                linked_user = User(
                    name=text_value("login_name") or franchise.franchisee_name or franchise.business_name,
                    surname=text_value("login_surname") or franchise.franchisee_surname or "User",
                    email=login_email, is_active=True, is_active_account=True,
                )
                linked_user.set_password("ChangeMe!2026")
                linked_user.roles.append(role)
                db.session.add(linked_user); db.session.flush()
                linked_user.assigned_franchises.append(franchise)
                db.session.flush()
            db.session.execute(user_franchises.update().where(
                user_franchises.c.franchise_id == franchise.id
            ).values(is_primary=False))
            db.session.execute(user_franchises.update().where(
                (user_franchises.c.franchise_id == franchise.id) &
                (user_franchises.c.user_id == linked_user.id)
            ).values(is_primary=True))
        RoyaltyScale.query.filter_by(franchise_id=franchise.id).delete(synchronize_session=False)
        for idx in range(1, 11):
            raw_from, raw_to, raw_pct = text_value(f"scale_{idx}_from"), text_value(f"scale_{idx}_to"), text_value(f"scale_{idx}_pct")
            if not any((raw_from, raw_to, raw_pct)): continue
            try:
                db.session.add(RoyaltyScale(franchise_id=franchise.id, row_number=idx,
                    amount_from=Decimal(raw_from or "0"), amount_to=Decimal(raw_to or "0"), percentage=Decimal(raw_pct or "0")))
            except InvalidOperation:
                db.session.rollback(); flash(f"Royalty scale row {idx} contains an invalid number.", "danger")
                return render_template("admin/edit_franchise_master_record.html", franchise=franchise, linked_user=linked_user, scales=scales)
        db.session.commit()
        log_action("Franchise Master", "Manual franchise update", f"Franchise {franchise.id}: {franchise.business_name}")
        flash("Franchise details saved to the database. The downloadable Master file now contains these changes.", "success")
        return redirect(url_for("admin.data_integrity"))
    return render_template("admin/edit_franchise_master_record.html", franchise=franchise, linked_user=linked_user, scales=scales)


@admin_bp.route("/data-integrity/franchise-master/export")
@admin_bp.route("/franchise-master/export")
@login_required
def export_franchise_master_update_template():
    """Download a populated Franchise Master workbook for correction and re-import."""
    from io import BytesIO
    from app.franchise_master_data import build_franchise_master_workbook

    if not (current_user.has_role("Admin") or current_user.has_role("Finance Manager") or current_user.has_role("Finance Assistant")):
        abort(403)
    data = build_franchise_master_workbook()
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Franchise_Master_Update_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )




@admin_bp.route("/data-integrity/needs-review/export")
@admin_bp.route("/franchise-master/needs-review/export")
@login_required
def export_needs_review_report():
    """Download a printable Excel report of franchise records that still need repair."""
    from io import BytesIO
    from app.franchise_master_data import build_needs_review_workbook

    if not (current_user.has_role("Admin") or current_user.has_role("Finance Manager") or current_user.has_role("Finance Assistant")):
        abort(403)
    data = build_needs_review_workbook()
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Franchise_Needs_Review_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    )

@admin_bp.route("/data-integrity/franchise-master/import", methods=["POST"])
@admin_bp.route("/franchise-master/import", methods=["POST"])
@login_required
def import_franchise_master_update_template():
    """Import the corrected Franchise Master workbook and update linked franchise records."""
    from app.franchise_master_data import import_franchise_master_workbook

    if not (current_user.has_role("Admin") or current_user.has_role("Finance Manager") or current_user.has_role("Finance Assistant")):
        abort(403)
    uploaded = request.files.get("excel_file")
    if not uploaded or not uploaded.filename:
        flash("Please choose the completed Franchise Master Excel file.", "warning")
        return redirect(url_for("admin.data_integrity"))
    try:
        result = import_franchise_master_workbook(uploaded)
        flash(f"Franchise Master import complete: {result['updated']} franchises updated, {result.get('changed_fields', 0)} fields changed, {result['scale_rows']} royalty scale rows updated, {len(result['unmatched'])} unmatched rows.", "success")
        if result["unmatched"]:
            flash("Some rows could not be matched. Download the report and check Franchise ID, Franchise Code or Business Name.", "warning")
    except Exception as exc:
        db.session.rollback()
        flash(f"Franchise Master import failed: {exc}", "danger")
    return redirect(url_for("admin.data_integrity"))
