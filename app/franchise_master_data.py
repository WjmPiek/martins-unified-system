"""Master Data Management helpers for franchise data integrity.

v104 makes the Franchise Master workbook the practical source of truth for
franchise details, region/province assignment, agreement dates and royalty
scales.  The importer is deliberately conservative: it updates existing
franchises only and matches by ID, then franchise code, then business name.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from app import db
from app.models import Franchise, MonthlyFigure, RoyaltyCalculationSnapshot, RoyaltyScale, User, user_franchises, ensure_mandatory_franchise_modules

PROVINCES = [
    "Eastern Cape", "Free State", "Gauteng", "KwaZulu-Natal", "Limpopo",
    "Mpumalanga", "North West", "Northern Cape", "Western Cape", "Unassigned",
]

ROYALTY_SCALE_COUNT = 10

MASTER_HEADERS = [
    "Franchise ID", "Readiness Status", "Issues To Fix", "Business Name", "Franchise Code",
    "Master Import ID", "Standardized Town", "Province", "Province Code", "Region", "District", "District Code", "Municipality", "Municipality Code", "Office Address", "Office Number",
    "After Hours Number", "Franchisee Name", "Franchisee Surname", "Franchisee Cell",
    "Franchisee Email", "Public Email", "Login User Email", "Agreement Start Date",
    "Agreement End Date", "Royalty Method", "Minimum Royalty Amount", "Active For Performance",
]
for idx in range(1, ROYALTY_SCALE_COUNT + 1):
    MASTER_HEADERS += [f"Scale {idx} From", f"Scale {idx} To", f"Scale {idx} %"]

ROYALTY_SCALE_HEADERS = [
    "Franchise ID", "Business Name", "Franchise Code", "Scale Row", "Amount From", "Amount To", "Percentage"
]

PROVINCE_TERMS = {
    "Gauteng": ["alberton", "benoni", "boksburg", "brakpan", "springs", "edenvale", "germiston", "katlehong", "vosloorus", "tsakane", "thokoza", "tokoza", "tembisa", "midrand", "pretoria", "soshanguve", "sochanguve", "mamelodi", "atteridgeville", "centurion", "hammanskraal", "vereeniging", "vanderbijlpark", "meyerton", "sebokeng", "orange farm", "lenasia", "three rivers", "florida", "fountainbleau", "fontainebleau", "carletonville", "randfontein", "krugersdorp", "roodepoort", "soweto", "sandton", "randburg", "johannesburg"],
    "Western Cape": ["cape town", "brackenfell", "bellville", "paarl", "parow", "kraaifontein", "kuils river", "durbanville", "table view", "stellenbosch", "strand", "somerset west", "worcester", "george", "mossel bay", "mosselbaai", "oudtshoorn", "knysna", "plettenberg", "beaufort west", "malmesbury", "vredenburg", "saldanha", "hermanus", "caledon", "robertson", "wellington"],
    "KwaZulu-Natal": ["durban", "pinetown", "phoenix", "umlazi", "umhlanga", "chatsworth", "isipingo", "kwamashu", "verulam", "tongaat", "pietermaritzburg", "empangeni", "richards bay", "ladysmith", "newcastle", "estcourt", "kokstad", "port shepstone", "margate", "vryheid", "eshowe", "ulundi", "stanger", "kwadukuza", "ballito"],
    "Eastern Cape": ["gqeberha", "port elizabeth", "east london", "mthatha", "umtata", "queenstown", "komani", "jeffreys bay", "jeffreys baai", "jeffreysbaai", "humansdorp", "uitenhage", "kariega", "grahamstown", "makhanda", "cradock", "graaff", "butterworth"],
    "Limpopo": ["polokwane", "pietersburg", "tzaneen", "mokopane", "potgietersrus", "mookgophong", "mookgopong", "modimolle", "nylstroom", "bela bela", "bela-bela", "belabela", "thohoyandou", "louis trichardt", "makhado", "giyani", "phalaborwa", "lephalale", "ellisras", "musina", "seshego"],
    "Mpumalanga": ["mbombela", "nelspruit", "witbank", "emalahleni", "middelburg", "secunda", "evander", "bethal", "ermelo", "piet retief", "barberton", "lydenburg", "white river", "hazyview", "komatipoort", "standerton", "volksrust", "delmas", "kriel"],
    "North West": ["rustenburg", "klerksdorp", "potchefstroom", "mahikeng", "mafikeng", "brits", "lichtenburg", "vryburg", "orkney", "stilfontein", "hartbeespoort", "zeerust", "taung", "wolmaransstad", "christiana"],
    "Free State": ["bloemfontein", "welkom", "bethlehem", "kroonstad", "sasolburg", "sasolsburg", "virginia", "harrismith", "parys", "ficksburg", "phuthaditjhaba", "botshabelo", "ladybrand", "senekal", "heilbron"],
    "Northern Cape": ["kimberley", "upington", "kuruman", "springbok", "de aar", "postmasburg", "kathu", "hartswater", "colesberg", "calvinia", "prieska", "douglas", "jan kempdorp", "barkly west", "warrenton", "hopetown"],
}


def normalize_key(value: Any) -> str:
    value = str(value or "").lower().strip()
    value = re.sub(r"\(f\)", "", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def infer_province(*values: Any) -> str:
    haystack = normalize_key(" ".join(str(v or "") for v in values))
    if not haystack:
        return ""
    for province, terms in PROVINCE_TERMS.items():
        for term in terms:
            if normalize_key(term) in haystack:
                return province
    return ""


def parse_date(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def decimal_or_zero(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, str):
        value = value.replace("R", "").replace("%", "").replace(",", "").strip()
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def bool_from_cell(value: Any, default: bool = True) -> bool:
    if value in (None, ""):
        return default
    text = str(value).strip().lower()
    if text in ("yes", "y", "true", "1", "active"):
        return True
    if text in ("no", "n", "false", "0", "inactive"):
        return False
    return default


def get_latest_period() -> Tuple[int, int]:
    row = db.session.query(MonthlyFigure.year, MonthlyFigure.month).order_by(MonthlyFigure.year.desc(), MonthlyFigure.month.desc()).first()
    if row:
        return int(row.year), int(row.month)
    today = date.today()
    return today.year, today.month


def franchise_lookup() -> Tuple[Dict[int, Franchise], Dict[str, Franchise], Dict[str, Franchise]]:
    by_id, by_code, by_name = {}, {}, {}
    for franchise in Franchise.query.all():
        by_id[franchise.id] = franchise
        if franchise.franchise_code:
            by_code[normalize_key(franchise.franchise_code)] = franchise
        by_name[normalize_key(franchise.business_name)] = franchise
    return by_id, by_code, by_name


def find_franchise(row: Dict[str, Any], lookups=None) -> Optional[Franchise]:
    by_id, by_code, by_name = lookups or franchise_lookup()

    # Franchise Code is the primary business key from v105 onward.
    code = normalize_key(row.get("Franchise Code"))
    if code and code in by_code:
        return by_code[code]

    master_import_id = normalize_key(row.get("Master Import ID"))
    standardized_town = normalize_key(row.get("Standardized Town"))
    if master_import_id or standardized_town:
        for franchise in Franchise.query.all():
            if master_import_id and normalize_key(getattr(franchise, "master_import_id", "")) == master_import_id:
                return franchise
            if standardized_town and normalize_key(getattr(franchise, "standardized_town", "")) == standardized_town:
                return franchise

    raw_id = row.get("Franchise ID")
    try:
        fid = int(raw_id) if raw_id not in (None, "") else None
    except Exception:
        fid = None
    if fid and fid in by_id:
        return by_id[fid]

    name = normalize_key(row.get("Business Name"))
    return by_name.get(name)


def franchise_login_email(franchise_id: int) -> str:
    user = (
        db.session.query(User)
        .join(user_franchises, User.id == user_franchises.c.user_id)
        .filter(user_franchises.c.franchise_id == franchise_id)
        .order_by(user_franchises.c.is_primary.desc(), User.id.asc())
        .first()
    )
    return user.email if user else ""


def readiness_issues(franchise: Franchise, latest_review=None) -> List[str]:
    issues = []
    if not franchise.business_name:
        issues.append("Business name missing")
    if not (getattr(franchise, "franchise_code", "") or "").strip():
        issues.append("Franchise code missing")
    if not (getattr(franchise, "province", "") or "").strip() or getattr(franchise, "province", "") == "Unassigned":
        issues.append("Province not assigned")
    if not (getattr(franchise, "region", "") or "").strip():
        issues.append("Region not assigned")
    if not franchise.office_address:
        issues.append("Office address missing")
    if not (franchise.office_number or franchise.after_hours_number or franchise.franchisee_cell):
        issues.append("Contact number missing")
    if not franchise.agreement_start_date:
        issues.append("Agreement start date missing")
    scale_count = RoyaltyScale.query.filter_by(franchise_id=franchise.id).count()
    if scale_count == 0:
        issues.append("Royalty scale missing")
    review_status = (getattr(latest_review, "status", "") or getattr(latest_review, "calculation_status", "") or "").lower()
    if latest_review and review_status in {"needs_review", "review", "failed", "error"}:
        reason = (
            getattr(latest_review, "reason", None)
            or getattr(latest_review, "diagnostic_message", None)
            or getattr(latest_review, "status_message", None)
            or getattr(latest_review, "message", None)
            or getattr(latest_review, "calculation_status", None)
            or "Royalty calculation needs review"
        )
        reason = str(reason).strip() or "Royalty calculation needs review"
        if reason not in issues:
            issues.append(reason)
    return issues



def ensure_franchise_codes(commit: bool = True) -> Dict[str, int]:
    """Assign stable MF### franchise codes to records that do not yet have one.

    The code is the permanent matching key used by Franchise Master and Month-End imports.
    Existing codes are preserved. New codes are allocated in business-name order and never
    re-used inside the same run.
    """
    existing = set()
    for f in Franchise.query.all():
        code = str(getattr(f, "franchise_code", "") or "").strip().upper()
        if code:
            existing.add(code)

    next_number = 1
    assigned = 0
    changed = 0
    for f in Franchise.query.order_by(Franchise.business_name, Franchise.id).all():
        code = str(getattr(f, "franchise_code", "") or "").strip().upper()
        if code:
            if getattr(f, "franchise_code", "") != code:
                f.franchise_code = code
                changed += 1
            continue
        while True:
            candidate = f"MF{next_number:03d}"
            next_number += 1
            if candidate not in existing:
                break
        f.franchise_code = candidate
        existing.add(candidate)
        assigned += 1
        changed += 1
    if commit and changed:
        db.session.commit()
    elif changed:
        db.session.flush()
    return {"assigned": assigned, "changed": changed, "total_codes": len(existing)}


def has_required_schema() -> Tuple[bool, List[str]]:
    """Lightweight database/schema guard used by the Data Integrity page."""
    required = ["franchise_code", "province", "region", "district", "municipality"]
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        columns = {c["name"] for c in inspector.get_columns("franchises")}
        missing = [name for name in required if name not in columns]
        return not missing, missing
    except Exception as exc:
        return False, [f"schema check failed: {exc}"]

def data_integrity_rows() -> List[Dict[str, Any]]:
    rows = []
    # Data Integrity is a master-data readiness screen. It must not mark a
    # franchise as Needs Review because of an old/stale royalty snapshot from
    # before the Master Import workbook was uploaded. Royalty import/calculation
    # issues are still handled in the royalty import and diagnostics pages.
    for f in Franchise.query.order_by(Franchise.business_name).all():
        scale_count = RoyaltyScale.query.filter_by(franchise_id=f.id).count()
        issues = readiness_issues(f, None)
        rows.append({
            "id": f.id,
            "business_name": f.business_name,
            "franchise_code": f.franchise_code or "",
            "province": getattr(f, "province", "") or "Unassigned",
            "region": getattr(f, "region", "") or "",
            "municipality": getattr(f, "municipality", "") or "",
            "agreement_start_date": f.agreement_start_date,
            "scale_count": scale_count,
            "login_email": franchise_login_email(f.id),
            "issue_count": len(issues),
            "issues": issues,
            "fix_hint": fix_hint(issues),
            "status": "Ready" if not issues else "Needs Review",
        })
    return rows


def fix_hint(issues: List[str]) -> str:
    if not issues:
        return "No action needed"
    hints = []
    for issue in issues:
        lower = issue.lower()
        if "franchise code" in lower or "code" in lower:
            hints.append("Keep or assign the permanent Franchise Code in the Franchise Master workbook.")
        elif "province" in lower:
            hints.append("Select the correct Province in the Franchise Master workbook.")
        elif "region" in lower:
            hints.append("Complete Region in the Franchise Master workbook.")
        elif "agreement" in lower:
            hints.append("Complete Agreement Start Date and Agreement Method.")
        elif "royalty scale" in lower or "scale" in lower:
            hints.append("Complete royalty bracket rows on the Franchise Master or Royalty Scales sheet.")
        elif "contact" in lower:
            hints.append("Complete Office Number, After Hours Number or Franchisee Cell.")
        elif "address" in lower:
            hints.append("Complete Office Address.")
    return " ".join(dict.fromkeys(hints)) or "Review and complete missing master data."


def assign_regions_from_existing_data(commit: bool = True) -> Dict[str, int]:
    updated = 0
    unassigned = 0
    for f in Franchise.query.order_by(Franchise.business_name).all():
        current = (getattr(f, "province", "") or "").strip()
        detected = infer_province(f.business_name, f.office_address, f.franchise_code, f.municipality, f.district, f.region)
        if detected:
            if current in ("", "Unassigned"):
                f.province = detected
                updated += 1
            if not getattr(f, "region", "") or getattr(f, "region", "") == "Unassigned":
                f.region = detected
        else:
            if not current:
                f.province = "Unassigned"
            unassigned += 1
    if commit:
        db.session.commit()
    return {"updated": updated, "unassigned": unassigned}


def build_franchise_master_workbook() -> bytes:
    wb = Workbook()
    # Remove default sheet after adding our structured sheets
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_master = wb.create_sheet("Franchise Master")
    ws_scales = wb.create_sheet("Royalty Scales")
    ws_report = wb.create_sheet("Needs Review Report")
    ws_lists = wb.create_sheet("Lists")
    ws_instructions = wb.create_sheet("Instructions")

    rows = data_integrity_rows()
    franchises = Franchise.query.order_by(Franchise.business_name).all()

    _write_summary(ws_summary, rows)
    _write_master(ws_master, franchises, rows)
    _write_scales(ws_scales, franchises)
    _write_report(ws_report, rows)
    _write_lists(ws_lists)
    _write_instructions(ws_instructions)
    _style_all_sheets(wb)
    _add_validations(wb, len(franchises) + 250)
    ws_lists.sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_needs_review_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Needs Review Report"
    rows = data_integrity_rows()
    _write_report(ws, rows)
    _style_all_sheets(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary(ws, rows):
    ready = sum(1 for r in rows if r["status"] == "Ready")
    review = len(rows) - ready
    no_province = sum(1 for r in rows if r["province"] == "Unassigned")
    no_scale = sum(1 for r in rows if r["scale_count"] == 0)
    latest_year, latest_month = get_latest_period()
    ws.append(["Franchise Master Workbook"])
    ws.append(["Latest reporting period", f"{latest_year}-{latest_month:02d}"])
    ws.append([])
    ws.append(["Metric", "Count"])
    ws.append(["Total Franchises", len(rows)])
    ws.append(["Ready", ready])
    ws.append(["Needs Review", review])
    ws.append(["No Province", no_province])
    ws.append(["No Royalty Scale", no_scale])
    ws.append([])
    ws.append(["Workflow"])
    ws.append(["1. Open the Needs Review Report sheet and fix each item in the Franchise Master sheet."])
    ws.append(["2. Complete Province and Region manually where the system cannot infer them."])
    ws.append(["3. Complete agreement dates and royalty brackets."])
    ws.append(["4. Import the updated file via Admin > Data Integrity."])


def _write_master(ws, franchises, integrity_rows):
    issues_by_id = {r["id"]: r for r in integrity_rows}
    ws.append(MASTER_HEADERS)
    for f in franchises:
        integrity = issues_by_id.get(f.id, {})
        scales = RoyaltyScale.query.filter_by(franchise_id=f.id).order_by(RoyaltyScale.row_number.asc()).all()
        row = [
            f.id,
            integrity.get("status", "Needs Review"),
            "; ".join(integrity.get("issues", [])),
            f.business_name,
            f.franchise_code,
            getattr(f, "master_import_id", "") or "",
            getattr(f, "standardized_town", "") or "",
            getattr(f, "province", "") or "Unassigned",
            getattr(f, "province_code", "") or "",
            getattr(f, "region", "") or "",
            getattr(f, "district", "") or "",
            getattr(f, "district_code", "") or "",
            getattr(f, "municipality", "") or "",
            getattr(f, "municipality_code", "") or "",
            f.office_address,
            f.office_number,
            f.after_hours_number,
            f.franchisee_name,
            f.franchisee_surname,
            f.franchisee_cell,
            f.franchisee_email,
            f.public_email,
            franchise_login_email(f.id),
            f.agreement_start_date,
            f.agreement_end_date,
            f.royalty_gross_method,
            "NONE" if getattr(f, "minimum_royalty_is_none", False) else float(f.minimum_royalty_amount or 0),
            "Yes" if getattr(f, "is_performance_active", True) else "No",
        ]
        for i in range(ROYALTY_SCALE_COUNT):
            s = scales[i] if i < len(scales) else None
            row.extend([float(s.amount_from or 0) if s else None, float(s.amount_to or 0) if s else None, float(s.percentage or 0) if s else None])
        ws.append(row)


def _write_scales(ws, franchises):
    ws.append(ROYALTY_SCALE_HEADERS)
    for f in franchises:
        scales = RoyaltyScale.query.filter_by(franchise_id=f.id).order_by(RoyaltyScale.row_number.asc()).all()
        if not scales:
            ws.append([f.id, f.business_name, f.franchise_code, 1, None, None, None])
            continue
        for s in scales:
            ws.append([f.id, f.business_name, f.franchise_code, s.row_number, float(s.amount_from or 0), float(s.amount_to or 0), float(s.percentage or 0)])


def _write_report(ws, rows):
    ws.append(["Franchise Data Integrity Fix Report"])
    latest_year, latest_month = get_latest_period()
    ws.append(["Latest reporting period", f"{latest_year}-{latest_month:02d}"])
    ws.append([])
    headers = ["ID", "Franchise Code", "Business Name", "Province", "Region", "Login Email", "Status", "Issues", "Suggested Fix"]
    ws.append(headers)
    for r in rows:
        if r["status"] != "Ready":
            ws.append([r["id"], r.get("franchise_code"), r["business_name"], r["province"], r["region"], r.get("login_email", ""), r["status"], "; ".join(r["issues"]), r["fix_hint"]])


def _write_lists(ws):
    ws.append(["Province", "Royalty Method", "YesNo"])
    max_len = max(len(PROVINCES), 2)
    for i in range(max_len):
        ws.append([
            PROVINCES[i] if i < len(PROVINCES) else None,
            ["old", "new"][i] if i < 2 else None,
            ["Yes", "No"][i] if i < 2 else None,
        ])


def _write_instructions(ws):
    lines = [
        ["Franchise Master Update Instructions"],
        ["This workbook is the easiest way to fix franchise master data in bulk."],
        ["Do not change the Franchise Code. The importer uses Franchise Code as the permanent primary match key."],
        ["Update Province, Region, Office Address, Contact Numbers, Agreement Dates and Royalty Scales."],
        ["Use yyyy-mm-dd for dates. Use numbers only for royalty amounts and percentages."],
        ["You may edit scale brackets either on the Franchise Master sheet or the Royalty Scales sheet."],
        ["The importer matches by Franchise Code first, then Franchise ID, then Business Name."],
        ["After import, run stabilize-platform or use the Admin rebuild buttons so dashboards and royalties refresh."],
    ]
    for row in lines:
        ws.append(row)


def _style_all_sheets(wb):
    header_fill = PatternFill("solid", fgColor="153D2A")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="153D2A")
    thin = Side(style="thin", color="D9E2D5")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.title in ("Summary", "Instructions", "Needs Review Report"):
            ws[1][0].font = title_font
        header_row = 1
        if ws.title in ("Summary", "Needs Review Report"):
            header_row = 4 if ws.title == "Needs Review Report" else 4
        if ws.title == "Instructions":
            ws.column_dimensions["A"].width = 110
            continue
        if ws.max_row >= header_row:
            for cell in ws[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            width = 16
            if col_idx in (3, 4, 8, 9, 10):
                width = 30
            if ws.title == "Franchise Master" and col_idx in (3, 10):
                width = 42
            if ws.title == "Needs Review Report" and col_idx in (8, 9):
                width = 50
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 24 if row_idx > 1 else 28
        if ws.title == "Franchise Master":
            header_map = _headers_for(ws)
            for header in ("Agreement Start Date", "Agreement End Date"):
                col = header_map.get(header)
                if col:
                    for col_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                        for c in col_cells:
                            c.number_format = "yyyy-mm-dd"
            number_headers = ["Minimum Royalty Amount"]
            for idx in range(1, ROYALTY_SCALE_COUNT + 1):
                number_headers.extend([f"Scale {idx} From", f"Scale {idx} To", f"Scale {idx} %"])
            for header in number_headers:
                col = header_map.get(header)
                if col:
                    for col_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                        for c in col_cells:
                            c.number_format = "#,##0.00"
        if ws.title == "Royalty Scales":
            for col in [5, 6, 7]:
                for col_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                    for c in col_cells:
                        c.number_format = "#,##0.00"


def _add_validations(wb, max_rows: int):
    master = wb["Franchise Master"]
    province_val = DataValidation(type="list", formula1="=Lists!$A$2:$A$11", allow_blank=True)
    method_val = DataValidation(type="list", formula1="=Lists!$B$2:$B$3", allow_blank=True)
    yes_no_val = DataValidation(type="list", formula1="=Lists!$C$2:$C$3", allow_blank=True)
    master.add_data_validation(province_val)
    master.add_data_validation(method_val)
    master.add_data_validation(yes_no_val)
    headers = _headers_for(master)
    province_col = get_column_letter(headers["Province"])
    method_col = get_column_letter(headers["Royalty Method"])
    active_col = get_column_letter(headers["Active For Performance"])
    province_val.add(f"{province_col}2:{province_col}{max_rows}")
    method_val.add(f"{method_col}2:{method_col}{max_rows}")
    yes_no_val.add(f"{active_col}2:{active_col}{max_rows}")


def _headers_for(ws) -> Dict[str, int]:
    return {str(c.value or "").strip(): idx + 1 for idx, c in enumerate(ws[1])}




# Columns accepted by the production master import file.  The system now accepts
# both the exported "Franchise Master" workbook and the polished single-tab
# "Master Import" workbook used as the operational source of truth.
MASTER_IMPORT_SHEET_NAMES = ("Master Import", "Franchise Master", "Master")
HEADER_ALIASES = {
    "Import ID": "Master Import ID",
    "Unique ID": "Master Import ID",
    "Master Import ID": "Master Import ID",
    "Town": "Standardized Town",
    "Standardized Town": "Standardized Town",
    "Business Name": "Business Name",
    "Franchise Name": "Business Name",
    "Franchise": "Business Name",
    "Franchise User": "Business Name",
    "Franchise Code": "Franchise Code",
    "Import ID": "Master Import ID",
    "Master Import ID": "Master Import ID",
    "Unique Import ID": "Master Import ID",
    "Unique ID": "Master Import ID",
    "Standardized Town": "Standardized Town",
    "Standardized Town Name": "Standardized Town",
    "Province": "Province",
    "Province Code": "Province Code",
    "District": "District",
    "District Municipality": "District",
    "District Code": "District Code",
    "Municipality": "Municipality",
    "Local/Metropolitan Municipality": "Municipality",
    "Local Municipality": "Municipality",
    "Municipality Code": "Municipality Code",
    "Municipal Code": "Municipality Code",
    "Region": "Region",
    "Office Address": "Office Address",
    "Office Number": "Office Number",
    "After Hours Number": "After Hours Number",
    "Franchisee Name": "Franchisee Name",
    "Franchisee Surname": "Franchisee Surname",
    "Franchisee Cell": "Franchisee Cell",
    "Franchisee Email": "Franchisee Email",
    "Public Email": "Public Email",
    "Login User Email": "Login User Email",
    "Agreement Start Date": "Agreement Start Date",
    "Agreement End Date": "Agreement End Date",
    "Royalty Method": "Royalty Method",
    "Minimum Royalty Amount": "Minimum Royalty Amount",
    "Minimum Royalty": "Minimum Royalty Amount",
    "Active For Performance": "Active For Performance",
    "Review Notes": "Review Notes",
}
for idx in range(1, ROYALTY_SCALE_COUNT + 1):
    HEADER_ALIASES[f"Scale {idx} From"] = f"Scale {idx} From"
    HEADER_ALIASES[f"Scale {idx} To"] = f"Scale {idx} To"
    HEADER_ALIASES[f"Scale {idx} %"] = f"Scale {idx} %"
    HEADER_ALIASES[f"Scale {idx} Percentage"] = f"Scale {idx} %"


def _clean_header_name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ").strip())


def _canonical_headers_for(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        raw = _clean_header_name(cell.value)
        canonical = HEADER_ALIASES.get(raw, raw)
        if canonical and canonical not in headers:
            headers[canonical] = idx
    return headers


def _select_master_worksheet(wb):
    for sheet_name in MASTER_IMPORT_SHEET_NAMES:
        if sheet_name in wb.sheetnames:
            return wb[sheet_name], sheet_name
    # Last fallback: use the active sheet when it has recognizable source-of-truth columns.
    ws = wb.active
    headers = _canonical_headers_for(ws)
    if {"Business Name", "Franchise Code"} & set(headers):
        return ws, ws.title
    raise ValueError("Workbook must contain a 'Master Import', 'Franchise Master' or 'Master' sheet.")


def _column_exists_on_franchise(field_name: str) -> bool:
    return hasattr(Franchise, field_name)


def _set_if_model_has(franchise: Franchise, field_name: str, value: Any) -> int:
    if not _column_exists_on_franchise(field_name):
        return 0
    if getattr(franchise, field_name) != value:
        setattr(franchise, field_name, value)
        return 1
    return 0


def _master_row_name(row: Dict[str, Any]) -> str:
    return str(row.get("Business Name") or row.get("Standardized Town") or row.get("Franchise Code") or "").strip()


def _master_row_town(row: Dict[str, Any]) -> str:
    return str(row.get("Standardized Town") or row.get("Business Name") or row.get("Franchise Code") or "").strip()


def _ensure_franchise_user_link(franchise: Franchise, row: Dict[str, Any]) -> Tuple[int, int]:
    """Create/link the franchise user from the Master Import row.

    Returns (users_created, users_linked).  Login User Email is preferred; if it
    is blank the importer uses Franchisee Email/Public Email, and if those are
    blank it creates the standard @martinsdirect.com login.  This makes the
    master import file the controlling allocation table before monthly/PDF imports.
    """
    login_email = str(row.get("Login User Email") or row.get("Franchisee Email") or row.get("Public Email") or "").strip().lower()
    if not login_email:
        slug = re.sub(r"[^a-z0-9]+", ".", _master_row_name(row).lower()).strip(".") or f"franchise{franchise.id}"
        login_email = f"{slug}@martinsdirect.com"
    user = User.query.filter(db.func.lower(User.email) == login_email.lower()).first()
    created = 0
    if not user:
        name = str(row.get("Franchisee Name") or _master_row_name(row) or "Franchise").strip()
        surname = str(row.get("Franchisee Surname") or "User").strip() or "User"
        user = User(name=name, surname=surname, email=login_email, is_active=True, is_active_account=True)
        try:
            user.set_password("ChangeMe!2026")
        except Exception:
            pass
        db.session.add(user)
        db.session.flush()
        created = 1
    role = None
    try:
        from app.models import Role
        role = Role.query.filter_by(name="Franchise User").first()
        if not role:
            role = Role(name="Franchise User", description="Imported franchise user role", is_system_role=True)
            db.session.add(role)
            db.session.flush()
        if role not in user.roles:
            user.roles.append(role)
        ensure_mandatory_franchise_modules(user)
    except Exception:
        pass
    linked = 0
    if franchise not in user.assigned_franchises:
        user.assigned_franchises.append(franchise)
        linked = 1
    return created, linked

def import_franchise_master_workbook(file_storage) -> Dict[str, Any]:
    wb = load_workbook(file_storage, data_only=True)
    ws, sheet_name = _select_master_worksheet(wb)
    headers = _canonical_headers_for(ws)
    required = ["Business Name", "Franchise Code"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))

    lookups = franchise_lookup()
    updated = 0
    created = 0
    users_created = 0
    users_linked = 0
    unmatched = []
    changed_fields = 0
    scale_updates = 0

    def cell(row_number, header):
        idx = headers.get(header)
        return ws.cell(row_number, idx).value if idx else None

    for r in range(2, ws.max_row + 1):
        row = {h: cell(r, h) for h in headers}
        if not any(row.get(h) for h in ("Franchise ID", "Business Name", "Franchise Code", "Standardized Town")):
            continue
        if str(row.get("Review Notes") or "").strip().lower() == "needs review":
            unmatched.append({"sheet": sheet_name, "row": r, "business_name": _master_row_name(row), "franchise_code": row.get("Franchise Code"), "reason": "Review Notes says Needs review"})
            continue
        franchise = find_franchise(row, lookups)
        if not franchise:
            # The Master Import file is now the source of truth.  Unknown rows are
            # created once they have a business name; the generated primary key is
            # then available for all later monthly/PDF imports.
            name = _master_row_name(row)
            if not name:
                unmatched.append({"sheet": sheet_name, "row": r, "business_name": row.get("Business Name"), "franchise_code": row.get("Franchise Code"), "reason": "No business name"})
                continue
            franchise = Franchise(business_name=name)
            db.session.add(franchise)
            db.session.flush()
            created += 1
            lookups = franchise_lookup()
        changed_fields += _update_franchise_from_row(franchise, row)
        # Refresh lookups after update because Franchise Code / Master Import ID
        # may have changed. Later duplicate rows must match the updated record.
        lookups = franchise_lookup()
        u_created, u_linked = _ensure_franchise_user_link(franchise, row)
        users_created += u_created
        users_linked += u_linked
        updated += 1
        parsed_scales = _scales_from_master_row(row)
        if parsed_scales:
            scale_updates += _replace_scales(franchise.id, parsed_scales)

    if "Royalty Scales" in wb.sheetnames:
        scale_updates += _import_scale_sheet(wb["Royalty Scales"], lookups, unmatched)

    db.session.commit()
    return {
        "updated": updated,
        "created": created,
        "changed_fields": changed_fields,
        "users_created": users_created,
        "users_linked": users_linked,
        "unmatched": unmatched,
        "scale_rows": scale_updates,
        "source_sheet": sheet_name,
    }

def _update_franchise_from_row(franchise: Franchise, row: Dict[str, Any]) -> int:
    changes = 0
    minimum_raw = row.get("Minimum Royalty Amount")
    minimum_is_none = str(minimum_raw or "").strip().lower() in {"none", "no minimum", "n/a"}
    province = str(row.get("Province") or infer_province(row.get("Business Name"), row.get("Office Address")) or "Unassigned").strip()
    updates = {
        "business_name": str(row.get("Business Name") or franchise.business_name or "").strip(),
        "franchise_code": str(row.get("Franchise Code") or franchise.franchise_code or "").strip(),
        "province": province,
        "region": str(row.get("Region") or province or "").strip(),
        "district": str(row.get("District") or "").strip(),
        "municipality": str(row.get("Municipality") or "").strip(),
        "office_address": str(row.get("Office Address") or "").strip(),
        "office_number": str(row.get("Office Number") or "").strip(),
        "after_hours_number": str(row.get("After Hours Number") or "").strip(),
        "franchisee_name": str(row.get("Franchisee Name") or "").strip(),
        "franchisee_surname": str(row.get("Franchisee Surname") or "").strip(),
        "franchisee_cell": str(row.get("Franchisee Cell") or "").strip(),
        "franchisee_email": str(row.get("Franchisee Email") or "").strip().lower(),
        "public_email": str(row.get("Public Email") or "").strip().lower(),
        "agreement_start_date": parse_date(row.get("Agreement Start Date")),
        "agreement_end_date": parse_date(row.get("Agreement End Date")),
        "minimum_royalty_amount": Decimal("0") if minimum_is_none else decimal_or_zero(minimum_raw),
        "minimum_royalty_is_none": minimum_is_none,
        "is_performance_active": bool_from_cell(row.get("Active For Performance"), getattr(franchise, "is_performance_active", True)),
    }
    method = str(row.get("Royalty Method") or franchise.royalty_gross_method or "old").lower().strip()
    updates["royalty_gross_method"] = method if method in ("old", "new") else "old"
    for field, value in updates.items():
        if getattr(franchise, field) != value:
            setattr(franchise, field, value)
            changes += 1
    changes += _set_if_model_has(franchise, "master_import_id", str(row.get("Master Import ID") or "").strip())
    changes += _set_if_model_has(franchise, "standardized_town", _master_row_town(row))
    changes += _set_if_model_has(franchise, "province_code", str(row.get("Province Code") or "").strip().upper())
    changes += _set_if_model_has(franchise, "district_code", str(row.get("District Code") or "").strip().upper())
    changes += _set_if_model_has(franchise, "municipality_code", str(row.get("Municipality Code") or "").strip().upper())
    return changes


def _scales_from_master_row(row: Dict[str, Any]) -> List[Tuple[int, Decimal, Decimal, Decimal]]:
    parsed = []
    for idx in range(1, ROYALTY_SCALE_COUNT + 1):
        amount_from = decimal_or_zero(row.get(f"Scale {idx} From"))
        amount_to = decimal_or_zero(row.get(f"Scale {idx} To"))
        percent = decimal_or_zero(row.get(f"Scale {idx} %"))
        if amount_from or amount_to or percent:
            parsed.append((idx, amount_from, amount_to, percent))
    return parsed


def _replace_scales(franchise_id: int, scales: List[Tuple[int, Decimal, Decimal, Decimal]]) -> int:
    RoyaltyScale.query.filter_by(franchise_id=franchise_id).delete()
    for idx, amount_from, amount_to, percent in scales:
        db.session.add(RoyaltyScale(franchise_id=franchise_id, row_number=idx, amount_from=amount_from, amount_to=amount_to, percentage=percent))
    return len(scales)


def _import_scale_sheet(ws, lookups, unmatched) -> int:
    headers = _headers_for(ws)
    if not all(h in headers for h in ROYALTY_SCALE_HEADERS):
        return 0
    grouped: Dict[int, List[Tuple[int, Decimal, Decimal, Decimal]]] = {}
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, headers[h]).value for h in headers}
        if not any(row.get(h) for h in ("Franchise ID", "Business Name", "Franchise Code")):
            continue
        franchise = find_franchise(row, lookups)
        if not franchise:
            unmatched.append({"sheet": "Royalty Scales", "row": r, "business_name": row.get("Business Name"), "franchise_code": row.get("Franchise Code")})
            continue
        scale_row = int(decimal_or_zero(row.get("Scale Row")) or 1)
        amount_from = decimal_or_zero(row.get("Amount From"))
        amount_to = decimal_or_zero(row.get("Amount To"))
        percent = decimal_or_zero(row.get("Percentage"))
        if amount_from or amount_to or percent:
            grouped.setdefault(franchise.id, []).append((scale_row, amount_from, amount_to, percent))
    count = 0
    for fid, scales in grouped.items():
        scales = sorted(scales, key=lambda item: item[0])
        count += _replace_scales(fid, scales)
    return count
