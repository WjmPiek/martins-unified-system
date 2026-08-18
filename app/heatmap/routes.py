from collections import Counter
from datetime import datetime
from functools import wraps
import math
from pathlib import Path

from flask import Blueprint, abort, current_app, jsonify, render_template, request, redirect, url_for, flash, send_file
from flask_login import current_user, login_required
from openpyxl import load_workbook

from app.audit import log_action
from app.extensions import db
from app.franchise_context import enforce_franchise_access, get_selected_franchise, is_franchise_view_mode
from app.models import Franchise, HeatmapRecord

heatmap_bp = Blueprint("heatmap", __name__, url_prefix="/heat-map")

PROVINCES = [
    "Eastern Cape", "Free State", "Gauteng", "KwaZulu-Natal", "Limpopo",
    "Mpumalanga", "North West", "Northern Cape", "Western Cape",
]

HEATMAP_RECORD_TYPES = {
    "deceased": "Deceased", "next_of_kin": "Next of Kin", "church": "Church", "cemetery": "Cemetery",
    "crematorium": "Crematorium", "insurance_clients": "Insurance Clients",
}

HEATMAP_TEMPLATE_FILES = {
    "insurance-members": (
        "insurance_members_heatmap_template.xlsx",
        "martins-insurance-members-heat-map-template.xlsx",
    ),
    "deceased-information": (
        "deceased_information_heatmap_template.xlsx",
        "martins-deceased-information-heat-map-template.xlsx",
    ),
}


def normalize_record_type(value):
    key = clean(value).lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "nok": "next_of_kin", "next_of_kin_client": "next_of_kin",
        "churches": "church", "cemeteries": "cemetery", "crematoria": "crematorium",
        "insurance": "insurance_clients", "insurance_client": "insurance_clients",
        "client": "insurance_clients", "clients": "insurance_clients",
    }
    key = aliases.get(key, key)
    return key if key in HEATMAP_RECORD_TYPES else "deceased"


def permission_required(code):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if not current_user.has_permission(code):
                abort(403)
            return func(*args, **kwargs)
        return wrapper
    return decorator


def can_modify_heatmap():
    return any(current_user.has_permission(code) for code in [
        "heat_map:add", "heat_map:edit", "heat_map:import", "heat_map:manage"
    ])


def can_import_heatmap():
    """Allow activated franchise users to import only into their own scope."""
    return current_user.has_permission("heat_map:import") or (
        current_user.is_franchise_scoped_user()
        and current_user.has_permission("heat_map:view")
    )


def accessible_franchises():
    return current_user.accessible_franchises()


def accessible_franchise_ids():
    return [franchise.id for franchise in accessible_franchises()]


def selected_or_requested_franchise_id():
    requested = request.values.get("franchise_id", type=int)
    selected = get_selected_franchise() if is_franchise_view_mode() else None
    franchise_id = requested or (selected.id if selected else None)
    allowed = set(accessible_franchise_ids())
    if franchise_id and franchise_id not in allowed:
        abort(403)
    return franchise_id


def import_target_franchise_id():
    """Resolve a compulsory, tenant-safe franchise target for an import."""
    allowed = accessible_franchise_ids()
    if current_user.is_franchise_scoped_user():
        # A franchise-side login may never import into a form/session-selected
        # branch outside its own scope.  Tenant isolation defines the target.
        return allowed[0] if len(allowed) == 1 else None
    return selected_or_requested_franchise_id()


def scoped_query():
    query = HeatmapRecord.query
    allowed = accessible_franchise_ids()
    if not (current_user.has_permission("franchise_management:view") or current_user.has_permission("franchise_management:manage")):
        if not allowed:
            return query.filter(False)
        query = query.filter(HeatmapRecord.franchise_id.in_(allowed))
    franchise_id = request.args.get("franchise_id", type=int)
    if franchise_id:
        if franchise_id not in allowed and not current_user.has_permission("franchise_management:view"):
            abort(403)
        query = query.filter_by(franchise_id=franchise_id)
    return query


def clean(value):
    return " ".join(str(value or "").strip().split())


def number(value):
    if value in (None, ""):
        return None
    try:
        parsed = float(str(value).strip().replace(",", "."))
        return parsed if math.isfinite(parsed) else None
    except ValueError:
        return None


def header_map(row):
    result = {}
    for index, value in enumerate(row, start=1):
        key = clean(value).lower().replace("_", " ").replace("-", " ")
        if key:
            result[key] = index
    return result


def cell(ws, row, headers, *names):
    for name in names:
        idx = headers.get(name)
        if idx:
            return ws.cell(row, idx).value
    return ""


def build_full_address(address, city, province, country, postal_code=""):
    location_parts = [clean(address), clean(city), clean(province), clean(postal_code)]
    cleaned_country = clean(country)
    if not any(location_parts) and not cleaned_country:
        return ""
    return ", ".join(part for part in [*location_parts, cleaned_country or "South Africa"] if part)


def header_positions(row):
    """Return every matching column so repeated Latitude/Longitude/Weight headers remain usable."""
    result = {}
    for index, value in enumerate(row, start=1):
        key = clean(value).lower().replace("_", " ").replace("-", " ")
        if key:
            result.setdefault(key, []).append(index)
    return result


def positioned_cell(ws, row, positions, name, occurrence=0):
    indexes = positions.get(name, [])
    if occurrence >= len(indexes):
        return ""
    return ws.cell(row, indexes[occurrence]).value


def parse_wide_density_template(ws, positions, source_filename, franchise_id):
    """Split one service row into deceased, venue and next-of-kin map records."""
    records = []
    skipped_blank = 0

    def value(row, name, occurrence=0):
        return clean(positioned_cell(ws, row, positions, name, occurrence))

    def numeric_value(row, name, occurrence=0):
        return number(positioned_cell(ws, row, positions, name, occurrence))

    def section_present(row, names, coordinate_occurrence):
        return (
            any(value(row, name) for name in names)
            or numeric_value(row, "latitude", coordinate_occurrence) is not None
            or numeric_value(row, "longitude", coordinate_occurrence) is not None
        )

    def add_record(*, row, record_type, name="", surname="", dod="", address="", city="",
                   province="", postal_code="", country="", full_address="", latitude=None,
                   longitude=None, weight=None, next_of_kin_name="", next_of_kin_surname="",
                   relationship="", contact_number="", mf_file="", enabled=True):
        if not enabled:
            return False
        meaningful = [
            mf_file, name, surname, dod, address, city, province, postal_code,
            full_address, next_of_kin_name, next_of_kin_surname, contact_number,
        ]
        if not any(clean(item) for item in meaningful) and latitude is None and longitude is None:
            return False
        resolved_full_address = full_address or build_full_address(address, city, province, country, postal_code)
        records.append(HeatmapRecord(
            franchise_id=franchise_id,
            mf_file=mf_file,
            deceased_name=name,
            deceased_surname=surname,
            dod=dod,
            address=address,
            city=city,
            province=province,
            country=country or "South Africa",
            full_address=resolved_full_address,
            latitude=latitude,
            longitude=longitude,
            weight=weight or 1,
            next_of_kin_name=next_of_kin_name,
            next_of_kin_surname=next_of_kin_surname,
            relationship=relationship,
            relation=f"MAP:{record_type}",
            contact_number=contact_number,
            source_filename=source_filename,
            created_by_id=current_user.id,
        ))
        return True

    for row in range(2, ws.max_row + 1):
        before = len(records)
        mf_file = value(row, "mf file")
        deceased_name = value(row, "deceased name")
        deceased_surname = value(row, "deceased surname")
        dod = value(row, "dod")
        nok_name = value(row, "next of kin name")
        nok_surname = value(row, "next of kin surname")
        relationship = value(row, "relationship")
        contact_number = value(row, "contact number")

        add_record(
            row=row, record_type="deceased", mf_file=mf_file, name=deceased_name,
            surname=deceased_surname, dod=dod, address=value(row, "address"),
            city=value(row, "city"), province=value(row, "province"),
            postal_code=value(row, "postal code"), country=value(row, "country"),
            full_address=value(row, "full address"), latitude=numeric_value(row, "latitude", 0),
            longitude=numeric_value(row, "longitude", 0), weight=numeric_value(row, "weight", 0),
            next_of_kin_name=nok_name, next_of_kin_surname=nok_surname,
            relationship=relationship, contact_number=contact_number,
            enabled=section_present(row, [
                "deceased name", "deceased surname", "dod", "address", "city",
                "province", "postal code", "full address",
            ], 0),
        )

        church_name = value(row, "church name")
        pastor_name = value(row, "pastor name")
        add_record(
            row=row, record_type="church", mf_file=mf_file, name=church_name, dod=dod,
            address=value(row, "church street address"), city=value(row, "church city"),
            province=value(row, "church province"), postal_code=value(row, "church postal code"),
            country=value(row, "church country"), full_address=value(row, "church full address"),
            latitude=numeric_value(row, "latitude", 1), longitude=numeric_value(row, "longitude", 1),
            weight=numeric_value(row, "weight", 1), next_of_kin_name=deceased_name,
            next_of_kin_surname=deceased_surname,
            relationship=f"Pastor: {pastor_name}" if pastor_name else "",
            enabled=section_present(row, [
                "church name", "church street address", "church city", "church province",
                "church postal code", "church full address", "pastor name",
            ], 1),
        )

        add_record(
            row=row, record_type="cemetery", mf_file=mf_file, name=value(row, "cemetery name"), dod=dod,
            address=value(row, "cemetery street address"), city=value(row, "cemetery city"),
            province=value(row, "cemetery province"), postal_code=value(row, "cemetery postal code"),
            country=value(row, "cemetery country"), full_address=value(row, "cemetery full address"),
            latitude=numeric_value(row, "latitude", 2), longitude=numeric_value(row, "longitude", 2),
            weight=numeric_value(row, "weight", 2), next_of_kin_name=deceased_name,
            next_of_kin_surname=deceased_surname,
            enabled=section_present(row, [
                "cemetery name", "cemetery street address", "cemetery city", "cemetery province",
                "cemetery postal code", "cemetery full address",
            ], 2),
        )

        add_record(
            row=row, record_type="crematorium", mf_file=mf_file, name=value(row, "crematorium name"), dod=dod,
            address=value(row, "crematorium street address"), city=value(row, "crematorium city"),
            province=value(row, "crematorium province"), postal_code=value(row, "crematorium postal code"),
            country=value(row, "crematorium country"), full_address=value(row, "crematorium full address"),
            latitude=numeric_value(row, "latitude", 3), longitude=numeric_value(row, "longitude", 3),
            weight=numeric_value(row, "weight", 3), next_of_kin_name=deceased_name,
            next_of_kin_surname=deceased_surname,
            enabled=section_present(row, [
                "crematorium name", "crematorium street address", "crematorium city",
                "crematorium province", "crematorium postal code",
                "crematorium full address",
            ], 3),
        )

        add_record(
            row=row, record_type="next_of_kin", mf_file=mf_file, name=deceased_name,
            surname=deceased_surname, dod=dod, address=value(row, "next of kin street address"),
            city=value(row, "next of kin city"), province=value(row, "next of kin province"),
            postal_code=value(row, "next of kin postal code"), country=value(row, "next of kin country"),
            full_address=value(row, "next of kin full address"), latitude=numeric_value(row, "latitude", 4),
            longitude=numeric_value(row, "longitude", 4), weight=numeric_value(row, "weight", 4),
            next_of_kin_name=nok_name, next_of_kin_surname=nok_surname,
            relationship=relationship, contact_number=contact_number,
            enabled=section_present(row, [
                "next of kin name", "next of kin surname", "relationship", "contact number",
                "next of kin street address", "next of kin city", "next of kin province",
                "next of kin postal code", "next of kin full address",
            ], 4),
        )

        if len(records) == before:
            skipped_blank += 1

    return records, 0, skipped_blank


def parse_heatmap_excel(file_storage, source_filename, franchise_id):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    header_values = [cell.value for cell in ws[1]]
    headers = header_map(header_values)
    if not headers:
        raise ValueError("The uploaded Excel file does not contain a header row.")

    positions = header_positions(header_values)
    wide_markers = {"church name", "cemetery name", "crematorium name", "next of kin street address"}
    if wide_markers.issubset(positions):
        return parse_wide_density_template(ws, positions, source_filename, franchise_id)

    has_relation_column = "relation" in headers
    records = []
    skipped_non_mem = 0
    skipped_blank = 0

    for row in range(2, ws.max_row + 1):
        mf_file = clean(cell(ws, row, headers, "mf file", "mf_file", "file", "policy number"))
        record_name = clean(cell(
            ws, row, headers, "record name", "location name", "church name",
            "cemetery name", "crematorium name",
        ))
        deceased_name = clean(cell(ws, row, headers, "deceased name", "name")) or record_name
        deceased_surname = clean(cell(ws, row, headers, "deceased surname", "surname"))
        address = clean(cell(ws, row, headers, "address", "street address", "residential address"))
        city = clean(cell(ws, row, headers, "city", "town", "town city"))
        province = clean(cell(ws, row, headers, "province"))
        country_value = clean(cell(ws, row, headers, "country"))
        country = country_value or "South Africa"
        full_address = clean(cell(ws, row, headers, "full address", "fulladdress")) or build_full_address(address, city, province, country_value)
        next_of_kin_name = clean(cell(ws, row, headers, "next of kin name", "nok name"))
        next_of_kin_surname = clean(cell(ws, row, headers, "next of kin surname", "nok surname"))
        relation = clean(cell(ws, row, headers, "relation"))
        explicit_record_type = clean(cell(ws, row, headers, "record type", "record_type", "map category", "category", "client type", "location type"))
        record_type = normalize_record_type(explicit_record_type)

        if has_relation_column and relation.upper() != "MEM":
            skipped_non_mem += 1
            continue
        if has_relation_column and relation.upper() == "MEM" and not explicit_record_type:
            record_type = "insurance_clients"
        if not any([
            mf_file, deceased_name, deceased_surname, next_of_kin_name,
            next_of_kin_surname, address, city, province, full_address,
        ]):
            skipped_blank += 1
            continue

        records.append(HeatmapRecord(
            franchise_id=franchise_id,
            mf_file=mf_file,
            deceased_name=deceased_name,
            deceased_surname=deceased_surname,
            dod=clean(cell(ws, row, headers, "dod", "date of death")),
            address=address,
            city=city,
            province=province,
            country=country,
            full_address=full_address,
            latitude=number(cell(ws, row, headers, "latitude", "lat")),
            longitude=number(cell(ws, row, headers, "longitude", "lng", "lon")),
            weight=number(cell(ws, row, headers, "weight")) or 1,
            next_of_kin_name=next_of_kin_name,
            next_of_kin_surname=next_of_kin_surname,
            relationship=clean(cell(ws, row, headers, "relationship")),
            relation=f"MAP:{record_type}" if explicit_record_type or relation.upper() == "MEM" else relation,
            contact_number=clean(cell(ws, row, headers, "contact number", "cell number", "phone")),
            source_filename=source_filename,
            created_by_id=current_user.id,
        ))
    return records, skipped_non_mem, skipped_blank


def heatmap_record_identity(record):
    """Build a stable per-franchise/category key for import reconciliation."""
    franchise_id = record.franchise_id
    record_type = record.map_record_type
    mf_file = clean(record.mf_file).casefold()
    if mf_file:
        return ("mf_file", franchise_id, record_type, mf_file)
    named_identity = (
        clean(record.deceased_name).casefold(),
        clean(record.deceased_surname).casefold(),
        clean(record.next_of_kin_name).casefold(),
        clean(record.next_of_kin_surname).casefold(),
        clean(record.dod).casefold(),
    )
    if any(named_identity):
        return ("details", franchise_id, record_type, *named_identity)
    return (
        "location", franchise_id, record_type,
        clean(record.full_address or record.address).casefold(),
        clean(record.city).casefold(),
        clean(record.province).casefold(),
        clean(record.contact_number).casefold(),
    )


IMPORT_MERGE_FIELDS = (
    "mf_file", "deceased_name", "deceased_surname", "dod", "address", "city",
    "province", "country", "full_address", "latitude", "longitude", "weight",
    "next_of_kin_name", "next_of_kin_surname", "relationship", "relation",
    "contact_number",
)


def merge_heatmap_record(existing, incoming):
    """Apply supplied values without erasing useful existing values with blanks."""
    changed = False
    for field in IMPORT_MERGE_FIELDS:
        incoming_value = getattr(incoming, field)
        if incoming_value is None or (isinstance(incoming_value, str) and not clean(incoming_value)):
            continue
        if getattr(existing, field) != incoming_value:
            setattr(existing, field, incoming_value)
            changed = True
    if changed:
        existing.source_filename = incoming.source_filename
    return changed


def reconcile_heatmap_records(records, franchise_id):
    """Synchronize an imported Heat Map snapshot for one franchise.

    Deceased-information and insurance-member workbooks are separate snapshots.
    Reimporting one removes stale imported locations from that same family while
    preserving the other family and locations created manually in the UI.
    """
    existing_records = HeatmapRecord.query.filter_by(franchise_id=franchise_id).all()
    by_identity = {}
    by_mf_file = {}
    duplicates_removed = 0
    for existing in existing_records:
        identity = heatmap_record_identity(existing)
        canonical = by_identity.get(identity)
        if canonical is not None:
            merge_heatmap_record(canonical, existing)
            db.session.delete(existing)
            duplicates_removed += 1
            continue
        by_identity[identity] = existing
        mf_file = clean(existing.mf_file).casefold()
        if mf_file:
            by_mf_file.setdefault(mf_file, []).append(existing)

    incoming_mf_counts = Counter(
        clean(record.mf_file).casefold()
        for record in records
        if clean(record.mf_file)
    )
    incoming_identities = {heatmap_record_identity(record) for record in records}
    incoming_types = {record.map_record_type for record in records}
    if incoming_types and incoming_types <= {"insurance_clients"}:
        synchronized_types = {"insurance_clients"}
    else:
        synchronized_types = {
            "deceased", "next_of_kin", "church", "cemetery", "crematorium"
        }

    additions = []
    updated = 0
    unchanged = 0
    for incoming in records:
        identity = heatmap_record_identity(incoming)
        existing = by_identity.get(identity)
        mf_file = clean(incoming.mf_file).casefold()
        # A narrow client file contains one row/category per MF File.  If a
        # legacy row has the same MF File but an old category marker, update it
        # instead of creating a second client.  Wide density files legitimately
        # contain several categories per MF File and therefore use exact keys.
        if existing is None and mf_file and incoming_mf_counts[mf_file] == 1:
            candidates = by_mf_file.get(mf_file, [])
            if len(candidates) == 1:
                existing = candidates[0]
        if existing is None:
            additions.append(incoming)
            by_identity[identity] = incoming
            if mf_file:
                by_mf_file.setdefault(mf_file, []).append(incoming)
        elif merge_heatmap_record(existing, incoming):
            updated += 1
            by_identity[identity] = existing
        else:
            unchanged += 1

    # Each template is the current snapshot for its family. Imported locations
    # absent from the new snapshot are stale. Manual rows have no source file
    # and are deliberately preserved.
    for existing in existing_records:
        if existing in db.session.deleted:
            continue
        if not clean(existing.source_filename):
            continue
        if existing.map_record_type not in synchronized_types:
            continue
        if heatmap_record_identity(existing) not in incoming_identities:
            db.session.delete(existing)
            duplicates_removed += 1

    if additions:
        db.session.add_all(additions)
    return len(additions), updated, unchanged, duplicates_removed


@heatmap_bp.route("/")
@login_required
@permission_required("heat_map:view")
def index():
    franchises = accessible_franchises()
    requested_id = request.args.get("franchise_id", type=int)
    selected = next((item for item in franchises if item.id == requested_id), None)
    if selected is None and is_franchise_view_mode():
        selected = get_selected_franchise()
    return render_template(
        "heatmap/index.html",
        franchises=franchises,
        selected_franchise=selected,
        provinces=PROVINCES,
        can_modify_heatmap=can_modify_heatmap(),
        can_import_heatmap=can_import_heatmap(),
        can_geocode_heatmap=can_modify_heatmap() or can_import_heatmap(),
        record_types=HEATMAP_RECORD_TYPES,
        google_maps_api_key=current_app.config.get("GOOGLE_MAPS_API_KEY", ""),
    )


@heatmap_bp.route("/template")
@heatmap_bp.route("/template/<template_type>")
@login_required
@permission_required("heat_map:view")
def download_template(template_type="deceased-information"):
    template_file = HEATMAP_TEMPLATE_FILES.get(template_type)
    if not template_file:
        abort(404)
    stored_name, download_name = template_file
    template_path = Path(current_app.static_folder) / "templates" / stored_name
    if not template_path.is_file():
        abort(404)
    return send_file(
        template_path,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@heatmap_bp.route("/data")
@login_required
@permission_required("heat_map:view")
def data():
    records = scoped_query().order_by(HeatmapRecord.city.asc(), HeatmapRecord.mf_file.asc()).all()
    serialized_records = [record.to_dict() for record in records]
    province_counts = Counter(record.province for record in records if record.province)
    city_counts = Counter(record.city for record in records if record.city)
    mapped = sum(
        1 for record in serialized_records
        if record["latitude"] is not None and record["longitude"] is not None
    )
    response = jsonify({
        "records": serialized_records,
        "summary": {
            "total": len(records),
            "mapped": mapped,
            "unmapped": len(records) - mapped,
            "province": dict(province_counts),
            "cities": dict(city_counts.most_common(10)),
        }
    })
    response.headers["Cache-Control"] = "no-store"
    return response


@heatmap_bp.route("/import", methods=["POST"])
@login_required
def import_excel():
    if not can_import_heatmap():
        abort(403)
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please choose an Excel file to import.", "danger")
        return redirect(url_for("heatmap.index"))
    filename = file.filename
    franchise_id = import_target_franchise_id()
    if not franchise_id:
        flash("Select an accessible franchise before importing Heat Map data.", "danger")
        return redirect(url_for("heatmap.index"))
    enforce_franchise_access(franchise_id)
    try:
        records, skipped_non_mem, skipped_blank = parse_heatmap_excel(file, filename, franchise_id)
        if not records:
            flash("No valid heat map rows were found in the uploaded file.", "warning")
            return redirect(url_for("heatmap.index"))
        added, updated, unchanged, removed = reconcile_heatmap_records(records, franchise_id)
        db.session.commit()
        detail = (
            f"Compared {len(records)} heat map records from {filename} for "
            f"{Franchise.query.get(franchise_id).business_name}. "
            f"Added: {added}; updated: {updated}; unchanged: {unchanged}. "
            f"Duplicate or stale imported locations removed: {removed}. "
            f"Skipped non-MEM rows: {skipped_non_mem}; blank rows: {skipped_blank}."
        )
        log_action("Heat Map", "Imported heat map records", detail)
        flash(detail, "success")
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Heat map import failed: %s", exc)
        flash(str(exc) or "Heat map import failed.", "danger")
    return redirect(url_for("heatmap.index", franchise_id=franchise_id))


@heatmap_bp.route("/record", methods=["POST"])
@login_required
def save_record():
    if not can_modify_heatmap():
        abort(403)
    payload = request.get_json(force=True)
    record_id = payload.get("id")
    record = HeatmapRecord.query.get(record_id) if record_id else HeatmapRecord(created_by_id=current_user.id)
    if not record:
        abort(404)
    if record.franchise_id:
        enforce_franchise_access(record.franchise_id)
    franchise_id = payload.get("franchiseId") or selected_or_requested_franchise_id()
    if franchise_id:
        franchise_id = int(franchise_id)
        enforce_franchise_access(franchise_id)
    record.franchise_id = franchise_id
    mapping = {
        "mf_file": "mfFile", "deceased_name": "deceasedName", "deceased_surname": "deceasedSurname",
        "dod": "dod", "address": "address", "city": "city", "province": "province", "country": "country",
        "full_address": "fullAddress", "next_of_kin_name": "nextOfKinName", "next_of_kin_surname": "nextOfKinSurname",
        "relationship": "relationship", "relation": "relation", "contact_number": "contactNumber",
    }
    for attr, key in mapping.items():
        setattr(record, attr, clean(payload.get(key)))
    if payload.get("recordType"):
        record.relation = f"MAP:{normalize_record_type(payload.get('recordType'))}"
    record.latitude = number(payload.get("latitude"))
    record.longitude = number(payload.get("longitude"))
    record.weight = number(payload.get("weight")) or 1
    if not record.full_address:
        record.full_address = build_full_address(record.address, record.city, record.province, record.country)
    db.session.add(record)
    db.session.commit()
    log_action("Heat Map", "Saved heat map record", record.mf_file or record.full_address or str(record.id))
    return jsonify({"record": record.to_dict()})


@heatmap_bp.route("/coordinates", methods=["POST"])
@login_required
def save_coordinates():
    """Save one geocoded address against all matching accessible records."""
    if not (can_modify_heatmap() or can_import_heatmap()):
        abort(403)
    payload = request.get_json(force=True) or {}
    raw_ids = payload.get("recordIds") or []
    try:
        record_ids = list(dict.fromkeys(int(value) for value in raw_ids))[:500]
    except (TypeError, ValueError):
        abort(400)
    latitude = number(payload.get("latitude"))
    longitude = number(payload.get("longitude"))
    if (
        not record_ids or latitude is None or longitude is None
        or not -90 <= latitude <= 90 or not -180 <= longitude <= 180
    ):
        abort(400)

    records = HeatmapRecord.query.filter(HeatmapRecord.id.in_(record_ids)).all()
    if len(records) != len(record_ids):
        abort(404)
    for record in records:
        enforce_franchise_access(record.franchise_id)
        record.latitude = latitude
        record.longitude = longitude
    db.session.commit()
    log_action(
        "Heat Map", "Geocoded heat map address",
        f"Updated {len(records)} location(s) at {records[0].full_address or records[0].address}",
    )
    return jsonify({"updated": len(records), "latitude": latitude, "longitude": longitude})


@heatmap_bp.route("/record/<int:record_id>/delete", methods=["POST"])
@login_required
def delete_record(record_id):
    if not can_modify_heatmap():
        abort(403)
    record = HeatmapRecord.query.get_or_404(record_id)
    enforce_franchise_access(record.franchise_id)
    db.session.delete(record)
    db.session.commit()
    log_action("Heat Map", "Deleted heat map record", record.mf_file or record.full_address or str(record_id))
    return jsonify({"message": "Deleted"})
